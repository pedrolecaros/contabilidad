"""Explorador de documentos: lista TODOS los archivos en backup por empresa
(libros SII, cartolas, F29, F22, fotos respaldo, otros). Permite descargar y eliminar."""
import os
from flask import (Blueprint, render_template, redirect, url_for, request, flash,
                   current_app, send_from_directory, abort)
from sqlalchemy import text
from models import db, Empresa, ArchivoImportado

bp = Blueprint('documentos', __name__)


def _rut_clean(rut: str) -> str:
    return (rut or '').replace('.', '').replace('-', '') or 'sin_rut'


def _scan_dir(root, rel='', max_depth=4):
    """Recorre el árbol y devuelve lista plana de archivos."""
    items = []
    full = os.path.join(root, rel) if rel else root
    if not os.path.isdir(full):
        return items
    try:
        entries = sorted(os.listdir(full))
    except (OSError, PermissionError):
        return items
    for name in entries:
        if name.startswith('.'):
            continue
        path = os.path.join(full, name)
        rel_path = os.path.join(rel, name) if rel else name
        if os.path.isdir(path):
            depth = rel_path.count(os.sep)
            if depth < max_depth:
                items.extend(_scan_dir(root, rel_path, max_depth))
        elif os.path.isfile(path):
            try:
                st = os.stat(path)
                parts = rel_path.split(os.sep)
                # parts puede ser [TIPO, periodo, filename] o [TIPO, filename] o solo [filename]
                tipo = parts[0] if len(parts) >= 2 else '(raíz)'
                periodo = parts[1] if len(parts) >= 3 else ''
                fname = parts[-1]
                items.append({
                    'rel_path': rel_path,
                    'nombre': fname,
                    'tipo': tipo,
                    'periodo': periodo if periodo not in ('sin_periodo', '') else '',
                    'tamano': st.st_size,
                    'mtime': st.st_mtime,
                })
            except OSError:
                pass
    return items


@bp.route('/consolidado/archivos/subir', methods=['POST'])
def subir_consolidado():
    """Sube un archivo desde el consolidado seleccionando empresa."""
    from storage import save_import_backup
    from flask import current_app as _ca
    from models import ArchivoImportado
    eid_str = request.form.get('empresa_id', '').strip()
    if not eid_str:
        flash('Seleccioná una empresa', 'warning')
        return redirect(url_for('documentos.consolidado'))
    try:
        eid = int(eid_str)
    except ValueError:
        flash('empresa_id inválido', 'danger')
        return redirect(url_for('documentos.consolidado'))
    empresa = Empresa.query.get(eid)
    if not empresa:
        flash('Empresa no existe', 'danger')
        return redirect(url_for('documentos.consolidado'))
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash('Seleccioná un archivo', 'warning')
        return redirect(url_for('documentos.consolidado'))
    periodo = (request.form.get('periodo') or '').strip() or None
    archivo.stream.seek(0)
    blob = archivo.stream.read()
    sub_rel = save_import_backup(blob, archivo.filename,
                                  _ca.config['UPLOAD_FOLDER'],
                                  empresa.rut, 'OTROS', periodo)
    import hashlib
    sha = hashlib.sha256(blob).hexdigest()
    registro = ArchivoImportado(
        empresa_id=eid, tipo='OTROS', nombre_archivo=archivo.filename,
        sha256=sha, periodo=periodo or '',
        fecha_importacion=__import__('datetime').datetime.now(), ndocs=0,
        respaldo_url=f'local:{sub_rel}',
    )
    db.session.add(registro); db.session.commit()
    flash(f'"{archivo.filename}" guardado en {empresa.razon_social} ({"global" if not periodo else periodo})', 'success')
    return redirect(url_for('documentos.consolidado'))


@bp.route('/consolidado/archivos')
def consolidado():
    """Explorador consolidado: archivos de TODAS las empresas activas con filtros."""
    empresas = Empresa.query.filter_by(activa=True).order_by(Empresa.razon_social).all()
    base_root = os.path.join(current_app.config['UPLOAD_FOLDER'], 'backups_importacion')

    archivos = []
    for emp in empresas:
        rut_clean = _rut_clean(emp.rut)
        emp_dir = os.path.join(base_root, rut_clean)
        for a in _scan_dir(emp_dir, max_depth=5):
            a['empresa_id'] = emp.id
            a['empresa_nombre'] = emp.nombre_fantasia or emp.razon_social
            a['empresa_rut'] = emp.rut
            archivos.append(a)

    # Filtros
    f_emp = request.args.get('empresa', '').strip()
    f_tipo = request.args.get('tipo', '').strip().upper()
    f_periodo = request.args.get('periodo', '').strip()
    f_search = request.args.get('q', '').strip().lower()

    if f_emp:
        archivos = [a for a in archivos if str(a['empresa_id']) == f_emp]
    if f_tipo:
        archivos = [a for a in archivos if a['tipo'].upper() == f_tipo]
    if f_periodo == '__global__':
        archivos = [a for a in archivos if not a['periodo']]
    elif f_periodo:
        archivos = [a for a in archivos if a['periodo'] == f_periodo]
    if f_search:
        archivos = [a for a in archivos if f_search in a['nombre'].lower()]

    archivos.sort(key=lambda x: x['mtime'], reverse=True)

    # Listas para filtros
    all_files = []
    for emp in empresas:
        rut_clean = _rut_clean(emp.rut)
        all_files.extend(_scan_dir(os.path.join(base_root, rut_clean), max_depth=5))
    tipos = sorted({a['tipo'] for a in all_files})
    periodos = sorted({a['periodo'] for a in all_files if a['periodo']}, reverse=True)
    total_bytes = sum(a['tamano'] for a in archivos)

    return render_template('documentos/consolidado.html',
        empresas=empresas, archivos=archivos,
        tipos=tipos, periodos=periodos,
        f_emp=f_emp, f_tipo=f_tipo, f_periodo=f_periodo, f_search=f_search,
        total_bytes=total_bytes, total_count=len(all_files))


@bp.route('/empresa/<int:eid>/archivos')
def index(eid):
    empresa = Empresa.query.get_or_404(eid)
    rut_clean = _rut_clean(empresa.rut)
    base = os.path.join(current_app.config['UPLOAD_FOLDER'],
                        'backups_importacion', rut_clean)
    archivos = _scan_dir(base, max_depth=5)

    # Filtros
    f_tipo = request.args.get('tipo', '').strip().upper()
    f_periodo = request.args.get('periodo', '').strip()
    f_search = request.args.get('q', '').strip().lower()

    if f_tipo:
        archivos = [a for a in archivos if a['tipo'].upper() == f_tipo]
    if f_periodo == '__global__':
        archivos = [a for a in archivos if not a['periodo']]
    elif f_periodo:
        archivos = [a for a in archivos if a['periodo'] == f_periodo]
    if f_search:
        archivos = [a for a in archivos if f_search in a['nombre'].lower()]

    # Ordenar por mtime desc
    archivos.sort(key=lambda x: x['mtime'], reverse=True)

    # Agregados: tipos y periodos únicos para los filtros
    all_archivos = _scan_dir(base, max_depth=5)
    tipos = sorted({a['tipo'] for a in all_archivos})
    periodos = sorted({a['periodo'] for a in all_archivos if a['periodo']}, reverse=True)

    total_bytes = sum(a['tamano'] for a in archivos)

    return render_template('documentos/index.html',
        empresa=empresa, archivos=archivos,
        tipos=tipos, periodos=periodos,
        f_tipo=f_tipo, f_periodo=f_periodo, f_search=f_search,
        total_bytes=total_bytes, total_count=len(all_archivos),
        rut_clean=rut_clean)


def _full_path(empresa, rel):
    """Resuelve la ruta completa validando que sea dentro del directorio de la empresa."""
    if not rel or '..' in rel.split('/'):
        return None
    rut_clean = _rut_clean(empresa.rut)
    base = os.path.join(current_app.config['UPLOAD_FOLDER'],
                        'backups_importacion', rut_clean)
    full = os.path.join(base, rel)
    if not os.path.isfile(full):
        return None
    return full


@bp.route('/empresa/<int:eid>/archivos/descargar')
def descargar(eid):
    empresa = Empresa.query.get_or_404(eid)
    rel = request.args.get('rel', '').strip()
    full = _full_path(empresa, rel)
    if not full:
        flash(f'Archivo no encontrado: {rel}', 'danger')
        return redirect(url_for('documentos.index', eid=eid))
    return send_from_directory(os.path.dirname(full), os.path.basename(full),
                                as_attachment=True)


@bp.route('/empresa/<int:eid>/archivos/inline')
def inline(eid):
    """Sirve el archivo inline (sin forzar download) — para PDF/imagen en <iframe>/<img>."""
    empresa = Empresa.query.get_or_404(eid)
    rel = request.args.get('rel', '').strip()
    full = _full_path(empresa, rel)
    if not full:
        abort(404)
    return send_from_directory(os.path.dirname(full), os.path.basename(full),
                                as_attachment=False)


@bp.route('/empresa/<int:eid>/archivos/preview')
def preview(eid):
    """Render preview HTML para CSV/XLS/XLSX. PDFs e imágenes van inline."""
    empresa = Empresa.query.get_or_404(eid)
    rel = request.args.get('rel', '').strip()
    full = _full_path(empresa, rel)
    if not full:
        abort(404)
    ext = os.path.splitext(full)[1].lower()
    fname = os.path.basename(full)

    rows = []
    error = None
    truncated = False
    LIMIT = 500

    try:
        if ext == '.csv':
            import csv as _csv
            with open(full, encoding='utf-8', errors='replace') as f:
                reader = _csv.reader(f, delimiter=';' if ';' in f.readline() else ',')
                f.seek(0)
                reader = _csv.reader(f, delimiter=';' if ';' in f.readline() else ',')
                f.seek(0)
                for i, row in enumerate(_csv.reader(f)):
                    if i >= LIMIT:
                        truncated = True; break
                    rows.append(row)
        elif ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(full, data_only=True, read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= LIMIT:
                    truncated = True; break
                rows.append(['' if v is None else str(v) for v in row])
        elif ext in ('.xls',):
            # XLS puede ser: (a) binario BIFF clásico, (b) HTML disfrazado (SII),
            # (c) CSV/texto con extensión .xls (cartolas algunos bancos).
            with open(full, 'rb') as f:
                head_raw = f.read(2048)
            head = head_raw.lstrip().lower()
            is_binary_xls = head_raw[:4] == b'\xd0\xcf\x11\xe0'
            is_html = (not is_binary_xls) and head.startswith(b'<') and (b'<table' in head or b'<html' in head)
            is_text = (not is_binary_xls) and (not is_html) and all(
                b in (9, 10, 13) or 32 <= b < 127 or b > 127 for b in head_raw[:512])
            if is_text and not is_html:
                # CSV/TSV disfrazado de .xls
                import csv as _csv
                with open(full, encoding='utf-8', errors='replace') as f:
                    sample = f.read(2048); f.seek(0)
                    delim = ';' if sample.count(';') > sample.count(',') else (',' if ',' in sample else '\t')
                    for i, row in enumerate(_csv.reader(f, delimiter=delim)):
                        if i >= LIMIT:
                            truncated = True; break
                        rows.append(row)
            elif is_html:
                # Parsear como HTML: extraer <tr><td> de la primera tabla
                from html.parser import HTMLParser
                class _TblParser(HTMLParser):
                    def __init__(self):
                        super().__init__()
                        self.rows = []; self.row = None; self.cell = None
                        self.in_table = False; self.cell_text = ''
                    def handle_starttag(self, tag, attrs):
                        t = tag.lower()
                        if t == 'table': self.in_table = True
                        elif t == 'tr' and self.in_table: self.row = []
                        elif t in ('td','th') and self.row is not None: self.cell = []; self.cell_text = ''
                        elif t == 'br' and self.cell is not None: self.cell_text += ' '
                    def handle_endtag(self, tag):
                        t = tag.lower()
                        if t in ('td','th') and self.cell is not None and self.row is not None:
                            self.row.append(self.cell_text.strip()); self.cell = None; self.cell_text = ''
                        elif t == 'tr' and self.row is not None:
                            self.rows.append(self.row); self.row = None
                        elif t == 'table': self.in_table = False
                    def handle_data(self, data):
                        if self.cell is not None: self.cell_text += data
                with open(full, encoding='utf-8', errors='replace') as f:
                    txt = f.read()
                p = _TblParser(); p.feed(txt)
                for r in p.rows[:LIMIT]:
                    rows.append(r)
                if len(p.rows) > LIMIT: truncated = True
            else:
                import xlrd
                wb = xlrd.open_workbook(full)
                sh = wb.sheet_by_index(0)
                for r in range(min(sh.nrows, LIMIT)):
                    rows.append([str(sh.cell_value(r, c)) for c in range(sh.ncols)])
                if sh.nrows > LIMIT: truncated = True
        elif ext in ('.txt', '.log'):
            with open(full, encoding='utf-8', errors='replace') as f:
                contenido = f.read(200_000)
            return render_template('documentos/preview.html',
                empresa=empresa, fname=fname, rel=rel,
                texto=contenido, truncated=len(contenido) >= 200_000)
        else:
            error = f'Tipo de archivo no previsualizable: {ext}. Usá Descargar.'
    except Exception as e:
        error = f'Error al leer archivo: {e}'

    return render_template('documentos/preview.html',
        empresa=empresa, fname=fname, rel=rel,
        rows=rows, truncated=truncated, error=error)


@bp.route('/empresa/<int:eid>/archivos/eliminar', methods=['POST'])
def eliminar(eid):
    empresa = Empresa.query.get_or_404(eid)
    rel = request.form.get('rel', '').strip()
    if not rel or '..' in rel.split('/'):
        flash('Ruta inválida', 'danger')
        return redirect(url_for('documentos.index', eid=eid))
    rut_clean = _rut_clean(empresa.rut)
    base = os.path.join(current_app.config['UPLOAD_FOLDER'],
                        'backups_importacion', rut_clean)
    full = os.path.join(base, rel)
    if not os.path.isfile(full):
        flash(f'Archivo no encontrado: {rel}', 'warning')
    else:
        try:
            os.remove(full)
            # Limpiar registro en archivos_importados si existe (matchea por sufijo)
            db.session.execute(text(
                "DELETE FROM archivos_importados WHERE empresa_id=:e AND respaldo_url LIKE :p"
            ), {'e': eid, 'p': f'%{rel}'})
            db.session.commit()
            flash(f'Archivo "{os.path.basename(rel)}" eliminado', 'info')
        except Exception as e:
            flash(f'Error al eliminar: {e}', 'danger')
    return redirect(url_for('documentos.index', eid=eid))

"""Carga movimientos Los Robles SpA 2026 desde Excel a Asientos en BORRADOR.

- Lee 'Movimientos Los Robles SpA 2026.xlsx'
- Reclasifica apertura: línea Futrono de 2.1.12 -> 2.1.11
- Genera asiento en BORRADOR (origen=BANCO) por cada fila
- Para enero: concilia con movimientos_banco existentes (Conciliacion tipo=MANUAL)
"""
import os
import sys
import sqlite3
from datetime import date

import openpyxl

DB = '/home/pedro/contabilidad/contabilidad.db'
XLSX = '/home/pedro/contabilidad/ejemplos/Contabilidades/04. Contable Los Robles SpA/Respaldos 2026/Movimientos Los Robles SpA 2026.xlsx'
EMP = 5  # Los Robles SpA

# Cuenta codigos -> id (Los Robles)
CUENTAS = {
    '1.1.02': 372, '1.1.06': 376, '1.1.09': 379, '1.1.12': 382, '1.1.14': 384,
    '2.1.05': 407, '2.1.06': 408, '2.1.07': 409, '2.1.11': 413, '2.1.12': 414,
    '2.1.13': 415,
    '4.1.02': 429, '4.2.03': 436,
    '5.2.01': 441, '5.2.02': 442, '5.2.07': 447, '5.2.09': 449, '5.2.10': 450,
    '5.2.11': 451, '5.2.13': 453, '5.2.16': 456, '5.2.17': 457,
}

# Contrapartes conocidas (id)
CP_ASESORIAS_ECOX = 3
CP_FELIPE_HIRIART = 5
CP_FUTRONO = 6
CP_BENJAMIN_LECAROS = 45
CP_FELIPE_CHAVEZ = 46
CP_SINDY_SANDOVAL = 49


def map_cuenta_contraparte(cat, det, com):
    """Devuelve (cuenta_codigo_contra, contraparte_id_para_cuenta_contra, descripcion_glosa)."""
    det_l = (det or '').lower()
    com_l = (com or '').lower()

    if cat == 'VENTA':
        return '4.1.02', None, None

    if cat == 'PRESTAMO':
        return '2.1.11', CP_FUTRONO, None

    if cat == 'FONDOS MUTUOS':
        return '1.1.09', None, None

    if cat == 'NOTARIA Y CONSERVADOR':
        if 'paga cbr' in com_l or 'walker' in det_l:
            return '4.2.03', None, None
        return '5.2.17', None, None

    if cat == 'IMPUESTOS':
        if 'dev impuesto' in det_l:
            return '4.2.03', None, None
        return '5.2.16', None, None

    if cat == 'OTRO':
        return '2.1.13', None, None

    if cat == 'SUELDO Y LEYES SOCIALES':
        if 'instituciones previsionales' in det_l or 'previred' in det_l or 'previred' in com_l:
            return '2.1.06', None, None
        # Sueldo Hector Varela directo
        return '5.2.01', None, None

    if cat == 'GASTOS GENERALES':
        if 'asesorias ecox' in det_l:
            if 'honorario' in com_l or 'dividendo' in com_l:
                return '5.2.02', CP_ASESORIAS_ECOX, None
            return '5.2.11', CP_ASESORIAS_ECOX, None
        if 'felipe andres hiriart' in det_l:
            if 'publicidad' in com_l:
                return '5.2.10', CP_FELIPE_HIRIART, None
            if 'uber' in com_l or 'reembolsa' in com_l:
                return '5.2.09', CP_FELIPE_HIRIART, None
            return '5.2.17', CP_FELIPE_HIRIART, None
        if 'hector varela' in det_l:
            return '5.2.17', None, None
        if 'sii.cl' in det_l:
            return '5.2.16', None, None
        if 'tesoreria' in det_l:
            return '5.2.16', None, None
        if 'servicio en internet' in det_l:
            return '5.2.16', None, None
        if 'sindy' in det_l:
            return '5.2.13', CP_SINDY_SANDOVAL, None
        if 'benjamin lecaros' in det_l:
            return '5.2.11', CP_BENJAMIN_LECAROS, None
        if 'la nube' in det_l:
            return '5.2.10', None, None
        if 'sandoval castro' in det_l or ('francisco' in det_l and 'sandoval' in det_l):
            return '5.2.07', None, None
        if 'troncoso' in det_l:
            return '5.2.17', None, None
        if 'abarca' in det_l:
            return '5.2.07', None, None
        if 'felipe chavez' in det_l:
            return '5.2.17', CP_FELIPE_CHAVEZ, None
        return '5.2.17', None, None

    # fallback
    return '5.2.17', None, None


def main():
    if not os.path.exists(DB):
        sys.exit(f'DB no existe: {DB}')
    if not os.path.exists(XLSX):
        sys.exit(f'Excel no existe: {XLSX}')

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['MOVIMIENTOS']

    conn = sqlite3.connect(DB)
    conn.execute('BEGIN')
    cur = conn.cursor()

    # 1) Reclasificar apertura: 2.1.12 (414) -> 2.1.11 (413) en línea Futrono
    cur.execute("""update lineas_asiento
                   set cuenta_id = ?, descripcion = 'Cuenta por Pagar Futrono SpA – desde 204101 (reclasificado 2.1.11)'
                   where asiento_id = 17 and cuenta_id = ?""",
                (CUENTAS['2.1.11'], CUENTAS['2.1.12']))
    print(f"Apertura reclasificada: {cur.rowcount} línea(s) 2.1.12->2.1.11")

    # 2) Cargar mov_banco enero (los 27)
    movs_jan = cur.execute("""select id, fecha, descripcion, cargo, abono
                              from movimientos_banco
                              where empresa_id = ? and fecha < '2026-02-01'
                              order by fecha, id""", (EMP,)).fetchall()
    print(f"Mov_banco enero disponibles: {len(movs_jan)}")
    movs_used = set()

    def buscar_mov(fecha_str, monto_dep, monto_giro):
        """Encuentra mov_banco matching fecha + cargo/abono."""
        for mid, mfecha, mdesc, mcargo, mabono in movs_jan:
            if mid in movs_used:
                continue
            if mfecha != fecha_str:
                continue
            if monto_dep and abs((mabono or 0) - monto_dep) < 1:
                return mid, mdesc
            if monto_giro and abs((mcargo or 0) - abs(monto_giro)) < 1:
                return mid, mdesc
        return None, None

    # 3) Numero siguiente
    next_num = (cur.execute("select coalesce(max(numero),0) from asientos where empresa_id=?", (EMP,)).fetchone()[0] or 0) + 1

    created_asientos = 0
    conciliated = 0
    sin_match_enero = []

    # 4) Iterar filas
    for r in range(3, ws.max_row + 1):
        fecha = ws.cell(r, 1).value
        det = ws.cell(r, 2).value
        cat = ws.cell(r, 3).value
        parcela = ws.cell(r, 5).value
        dep = ws.cell(r, 7).value
        giro = ws.cell(r, 8).value
        com = ws.cell(r, 9).value
        if not fecha:
            continue
        if not (dep or giro):
            continue

        fecha_str = fecha.strftime('%Y-%m-%d')
        es_ingreso = bool(dep)
        monto = float(dep) if es_ingreso else abs(float(giro))

        cuenta_contra_cod, cp_id, _ = map_cuenta_contraparte(cat, det, com)
        cuenta_contra_id = CUENTAS[cuenta_contra_cod]
        cuenta_banco_id = CUENTAS['1.1.02']

        # Descripción del asiento
        partes = []
        if cat == 'VENTA':
            partes.append('Venta lote')
            if com:
                partes.append(com)
            if parcela:
                partes.append(f'({parcela})')
        elif cat == 'PRESTAMO':
            partes.append('Abono préstamo Futrono SpA')
        elif cat == 'FONDOS MUTUOS':
            partes.append('Rescate FFMM' if es_ingreso else 'Aporte FFMM')
        elif cat == 'SUELDO Y LEYES SOCIALES':
            partes.append('Pago Previred' if cuenta_contra_cod == '2.1.06' else 'Sueldo Hector Varela')
        elif cat == 'NOTARIA Y CONSERVADOR':
            if cuenta_contra_cod == '4.2.03':
                partes.append('Recupera CBR')
            else:
                partes.append('Notaría/CBR')
            if det:
                partes.append(det)
        elif cat == 'IMPUESTOS':
            if cuenta_contra_cod == '4.2.03':
                partes.append('Devolución impuesto')
            else:
                partes.append('Pago impuesto')
            if com:
                partes.append(com)
        elif cat == 'OTRO':
            partes.append(com or det or 'Otro movimiento')
        else:  # GASTOS GENERALES
            if det:
                partes.append(det)
            if com:
                partes.append(f'– {com}')
        descripcion = ' '.join(str(p) for p in partes if p)[:480]

        # Crear asiento BORRADOR
        cur.execute("""insert into asientos
                       (empresa_id, fecha, numero, descripcion, origen, estado, creado_en)
                       values (?, ?, ?, ?, 'BANCO', 'BORRADOR', datetime('now'))""",
                    (EMP, fecha_str, next_num, descripcion))
        asiento_id = cur.lastrowid
        next_num += 1
        created_asientos += 1

        # Líneas (debe / haber)
        glosa_linea = (com or det or '')[:280]

        if es_ingreso:
            # Banco debe / contra haber
            cur.execute("""insert into lineas_asiento
                           (asiento_id, cuenta_id, debe, haber, descripcion, orden, contraparte_id)
                           values (?, ?, ?, 0, ?, 1, NULL)""",
                        (asiento_id, cuenta_banco_id, monto, glosa_linea))
            cur.execute("""insert into lineas_asiento
                           (asiento_id, cuenta_id, debe, haber, descripcion, orden, contraparte_id)
                           values (?, ?, 0, ?, ?, 2, ?)""",
                        (asiento_id, cuenta_contra_id, monto, glosa_linea, cp_id))
        else:
            # Contra debe / Banco haber
            cur.execute("""insert into lineas_asiento
                           (asiento_id, cuenta_id, debe, haber, descripcion, orden, contraparte_id)
                           values (?, ?, ?, 0, ?, 1, ?)""",
                        (asiento_id, cuenta_contra_id, monto, glosa_linea, cp_id))
            cur.execute("""insert into lineas_asiento
                           (asiento_id, cuenta_id, debe, haber, descripcion, orden, contraparte_id)
                           values (?, ?, 0, ?, ?, 2, NULL)""",
                        (asiento_id, cuenta_banco_id, monto, glosa_linea))

        # Audit CREAR
        cur.execute("""insert into asientos_audit (asiento_id, accion, descripcion, creado_en)
                       values (?, 'CREAR', ?, datetime('now'))""",
                    (asiento_id, f'Carga masiva desde Excel Movimientos LR 2026 fila {r}'))

        # 5) Concilia si es enero
        if fecha.month == 1 and fecha.year == 2026:
            mid, mdesc = buscar_mov(fecha_str, dep, giro)
            if mid:
                movs_used.add(mid)
                # Crear conciliación
                conc_desc = f'{descripcion[:100]} | {(mdesc or "")[:80]}'
                cur.execute("""insert into conciliaciones
                               (empresa_id, fecha, descripcion, tipo, respaldo_url, contraparte_id)
                               values (?, ?, ?, 'MANUAL', NULL, ?)""",
                            (EMP, fecha_str, conc_desc[:280], cp_id))
                conc_id = cur.lastrowid
                # Linkear movimiento_banco
                cur.execute("""update movimientos_banco
                               set asiento_id = ?, conciliacion_id = ?, procesado = 1
                               where id = ?""",
                            (asiento_id, conc_id, mid))
                conciliated += 1
            else:
                sin_match_enero.append((fecha_str, det, dep or giro))

    conn.commit()
    conn.close()

    print(f'\nAsientos creados: {created_asientos}')
    print(f'Conciliaciones enero creadas: {conciliated}')
    print(f'Mov_banco enero sin conciliar: {len(movs_jan) - len(movs_used)}')
    if sin_match_enero:
        print('\nFilas enero sin match en banco:')
        for x in sin_match_enero:
            print(f'  {x}')
    movs_pendientes = [m for m in movs_jan if m[0] not in movs_used]
    if movs_pendientes:
        print('\nMov_banco enero no conciliados:')
        for m in movs_pendientes:
            print(f'  {m}')


if __name__ == '__main__':
    main()

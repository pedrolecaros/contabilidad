from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, current_app, send_file
from models import db, Empresa, Empleado, Liquidacion, VariablesMensuales, ValorUF
from engine import remuneraciones as motor

bp = Blueprint('remuneraciones', __name__)


def _dias_habiles_entre(fecha_inicio, fecha_fin):
    """Cuenta días hábiles (lunes–viernes) entre dos fechas, inclusive."""
    from datetime import timedelta
    if fecha_fin < fecha_inicio:
        return 0
    dias = 0
    d = fecha_inicio
    while d <= fecha_fin:
        if d.weekday() < 5:  # 0=Mon … 4=Fri
            dias += 1
        d += timedelta(days=1)
    return dias

AFP_OPCIONES = ['Capital', 'Cuprum', 'Habitat', 'Modelo', 'PlanVital', 'ProVida', 'Uno']


@bp.route('/empresa/<int:eid>/remuneraciones')
def index(eid):
    empresa = Empresa.query.get_or_404(eid)
    empleados_activos = Empleado.query.filter_by(empresa_id=eid, activo=True).order_by(Empleado.nombre).all()
    empleados_inactivos = Empleado.query.filter_by(empresa_id=eid, activo=False).order_by(Empleado.nombre).all()
    return render_template('remuneraciones/index.html', empresa=empresa,
                           empleados=empleados_activos, empleados_inactivos=empleados_inactivos)


@bp.route('/empresa/<int:eid>/remuneraciones/nuevo', methods=['GET', 'POST'])
def nuevo(eid):
    empresa = Empresa.query.get_or_404(eid)
    if request.method == 'POST':
        emp = Empleado(empresa_id=eid)
        _poblar(emp, request.form)
        db.session.add(emp)
        db.session.commit()
        flash('Empleado creado correctamente.', 'success')
        return redirect(url_for('remuneraciones.index', eid=eid))
    return render_template('remuneraciones/form.html', empresa=empresa,
                           emp=None, afp_opciones=AFP_OPCIONES)


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/editar', methods=['GET', 'POST'])
def editar(eid, emp_id):
    empresa = Empresa.query.get_or_404(eid)
    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()
    if request.method == 'POST':
        _poblar(emp, request.form)
        db.session.commit()
        flash('Empleado actualizado.', 'success')
        return redirect(url_for('remuneraciones.index', eid=eid))
    return render_template('remuneraciones/form.html', empresa=empresa,
                           emp=emp, afp_opciones=AFP_OPCIONES)


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/eliminar', methods=['POST'])
def eliminar(eid, emp_id):
    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()
    emp.activo = False
    db.session.commit()
    flash(f'{emp.nombre} marcado como inactivo. Puedes reactivarlo desde la lista.', 'warning')
    return redirect(url_for('remuneraciones.index', eid=eid))


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/reactivar', methods=['POST'])
def reactivar(eid, emp_id):
    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()
    emp.activo = True
    db.session.commit()
    flash(f'{emp.nombre} reactivado.', 'success')
    return redirect(url_for('remuneraciones.index', eid=eid))


def _mes_anterior():
    from datetime import date
    hoy = date.today()
    if hoy.month == 1:
        return f'{hoy.year - 1}-12'
    return f'{hoy.year}-{hoy.month - 1:02d}'


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/liquidar', methods=['GET', 'POST'])
def liquidar(eid, emp_id):
    from datetime import date
    empresa = Empresa.query.get_or_404(eid)
    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()

    if request.method == 'POST':
        periodo = request.form.get('periodo', '').strip()
        horas_extra = float(request.form.get('horas_extra', 0) or 0)
        otros = float(request.form.get('otros', 0) or 0)
        dias_trabajados = max(1, min(30, int(request.form.get('dias_trabajados', 30) or 30)))
        accion = request.form.get('accion', 'calcular')

        if not periodo:
            flash('El período es obligatorio.', 'danger')
            periodo_default = _mes_anterior()
            vars_mes = VariablesMensuales.query.filter_by(periodo=periodo_default).first()
            return render_template('remuneraciones/liquidar.html', empresa=empresa, emp=emp,
                                   vars_mes=vars_mes, periodo_default=periodo_default,
                                   dias_trabajados=30)

        vars = VariablesMensuales.query.filter_by(periodo=periodo).first()
        if not vars:
            flash(f'No hay variables para {periodo}. Carga en Remuneraciones → Variables.', 'warning')
            periodo_default = _mes_anterior()
            vars_mes = VariablesMensuales.query.filter_by(periodo=periodo_default).first()
            return render_template('remuneraciones/liquidar.html', empresa=empresa, emp=emp,
                                   vars_mes=vars_mes, periodo_default=periodo_default,
                                   dias_trabajados=dias_trabajados)

        # Prorate if not full month
        from types import SimpleNamespace
        factor = dias_trabajados / 30.0
        if factor != 1.0:
            emp_calc = SimpleNamespace(
                sueldo_base=round(emp.sueldo_base * factor),
                bono_colacion=round((getattr(emp, 'bono_colacion', 0) or 0) * factor),
                bono_movilizacion=round((getattr(emp, 'bono_movilizacion', 0) or 0) * factor),
                otros_haberes=getattr(emp, 'otros_haberes', 0) or 0,
                afp=emp.afp, tasa_afp_comision=emp.tasa_afp_comision,
                tipo_salud=emp.tipo_salud, isapre=emp.isapre,
                monto_isapre=getattr(emp, 'monto_isapre', 0) or 0,
                monto_isapre_uf=getattr(emp, 'monto_isapre_uf', 0) or 0,
                tasa_mutual=getattr(emp, 'tasa_mutual', 0.0093) or 0.0093,
                tipo_sueldo=getattr(emp, 'tipo_sueldo', 'BRUTO'),
            )
        else:
            emp_calc = emp

        resultado = motor.calcular(
            emp_calc,
            utm=vars.utm,
            uf=vars.uf,
            tope_gratificacion=vars.tope_gratificacion,
            tope_imponible=vars.tope_imponible,
            horas_extra=horas_extra,
            otros=otros,
            tasa_sis=vars.tasa_sis or None,
        )

        if accion == 'calcular':
            return render_template('remuneraciones/liquidar.html', empresa=empresa, emp=emp,
                                   vars_mes=vars, periodo_default=periodo,
                                   preview=resultado, dias_trabajados=dias_trabajados,
                                   form_data={'periodo': periodo, 'horas_extra': horas_extra,
                                              'otros': otros, 'dias_trabajados': dias_trabajados})

        # accion == 'borrador' o 'emitir' → guardar
        existe = Liquidacion.query.filter_by(empleado_id=emp_id, periodo=periodo).first()
        if existe:
            flash(f'Ya existe una liquidación para {periodo}. Edítela o elimínela primero.', 'warning')
            return redirect(url_for('remuneraciones.detalle', eid=eid, liq_id=existe.id))

        estado = 'EMITIDA' if accion == 'emitir' else 'BORRADOR'
        liq = Liquidacion(empresa_id=eid, empleado_id=emp_id, periodo=periodo, estado=estado)
        for campo, valor in resultado.items():
            if hasattr(liq, campo):
                setattr(liq, campo, valor)
        db.session.add(liq)
        db.session.commit()
        if accion == 'emitir':
            # Generar PDF y guardarlo
            try:
                from services.pdf import guardar_liquidacion_pdf
                from flask import current_app
                storage_key = guardar_liquidacion_pdf(
                    current_app._get_current_object(), liq,
                    current_app.config['UPLOAD_FOLDER']
                )
                liq.archivo_url = storage_key
                db.session.commit()
            except Exception as e:
                flash(f'Liquidación emitida pero no se pudo generar el PDF: {e}', 'warning')
            return redirect(url_for('remuneraciones.imprimir', eid=eid, liq_id=liq.id))
        flash(f'Borrador {periodo} guardado.', 'success')
        return redirect(url_for('remuneraciones.detalle', eid=eid, liq_id=liq.id))

    # GET
    periodo_default = _mes_anterior()
    vars_mes = VariablesMensuales.query.filter_by(periodo=periodo_default).first()
    ultima_liq = (Liquidacion.query
                  .filter_by(empresa_id=eid, empleado_id=emp_id)
                  .order_by(Liquidacion.periodo.desc())
                  .first())
    return render_template('remuneraciones/liquidar.html', empresa=empresa, emp=emp,
                           vars_mes=vars_mes, periodo_default=periodo_default,
                           dias_trabajados=30, ultima_liq=ultima_liq)


@bp.route('/empresa/<int:eid>/remuneraciones/liquidacion/<int:liq_id>')
def detalle(eid, liq_id):
    empresa = Empresa.query.get_or_404(eid)
    liq = Liquidacion.query.filter_by(id=liq_id, empresa_id=eid).first_or_404()
    return render_template('remuneraciones/detalle.html', empresa=empresa, liq=liq)


@bp.route('/empresa/<int:eid>/remuneraciones/liquidacion/<int:liq_id>/imprimir')
def imprimir(eid, liq_id):
    empresa = Empresa.query.get_or_404(eid)
    liq = Liquidacion.query.filter_by(id=liq_id, empresa_id=eid).first_or_404()
    return render_template('remuneraciones/imprimir.html', empresa=empresa, liq=liq)


@bp.route('/empresa/<int:eid>/remuneraciones/liquidacion/<int:liq_id>/emitir', methods=['POST'])
def emitir_liq(eid, liq_id):
    from flask import current_app
    liq = Liquidacion.query.filter_by(id=liq_id, empresa_id=eid).first_or_404()
    liq.estado = 'EMITIDA'
    db.session.commit()
    if not liq.archivo_url:
        try:
            from services.pdf import guardar_liquidacion_pdf
            storage_key = guardar_liquidacion_pdf(
                current_app._get_current_object(), liq,
                current_app.config['UPLOAD_FOLDER']
            )
            liq.archivo_url = storage_key
            db.session.commit()
        except Exception as e:
            flash(f'Liquidación emitida pero no se pudo generar el PDF: {e}', 'warning')
    flash('Liquidación emitida.', 'success')
    return redirect(url_for('remuneraciones.detalle', eid=eid, liq_id=liq_id))


@bp.route('/empresa/<int:eid>/remuneraciones/liquidacion/<int:liq_id>/eliminar', methods=['POST'])
def eliminar_liq(eid, liq_id):
    liq = Liquidacion.query.filter_by(id=liq_id, empresa_id=eid).first_or_404()
    emp_id = liq.empleado_id
    db.session.delete(liq)
    db.session.commit()
    flash('Liquidación eliminada.', 'success')
    return redirect(url_for('remuneraciones.historial', eid=eid, emp_id=emp_id))


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/historial')
def historial(eid, emp_id):
    empresa = Empresa.query.get_or_404(eid)
    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()
    liqs = (Liquidacion.query
            .filter_by(empleado_id=emp_id, empresa_id=eid)
            .order_by(Liquidacion.periodo.desc())
            .all())
    return render_template('remuneraciones/historial.html', empresa=empresa, emp=emp, liqs=liqs)


@bp.route('/empresa/<int:eid>/remuneraciones/informe-renta')
def informe_renta(eid):
    from datetime import date
    empresa = Empresa.query.get_or_404(eid)
    hoy = date.today()
    anio = request.args.get('anio', hoy.year, type=int)

    # Query all EMITIDA liquidaciones for this year grouped by employee
    liqs = (Liquidacion.query
            .join(Empleado, Liquidacion.empleado_id == Empleado.id)
            .filter(
                Liquidacion.empresa_id == eid,
                Liquidacion.estado == 'EMITIDA',
                Liquidacion.periodo.like(f'{anio}-%'),
            )
            .order_by(Empleado.nombre, Liquidacion.periodo)
            .all())

    # Group by employee
    from collections import defaultdict
    grupos = {}  # emp_id -> dict
    detalles = defaultdict(list)  # emp_id -> [liq, ...]

    for liq in liqs:
        eid_emp = liq.empleado_id
        detalles[eid_emp].append(liq)
        if eid_emp not in grupos:
            grupos[eid_emp] = {
                'empleado': liq.empleado,
                'renta_acumulada': 0.0,
                'impuesto_acumulado': 0.0,
                'meses': 0,
            }
        grupos[eid_emp]['renta_acumulada'] += liq.renta_imponible or 0
        grupos[eid_emp]['impuesto_acumulado'] += liq.impuesto_renta or 0
        grupos[eid_emp]['meses'] += 1

    filas = []
    for eid_emp, g in grupos.items():
        meses = g['meses']
        impuesto_acum = g['impuesto_acumulado']
        proyeccion = round(impuesto_acum * 12 / meses) if meses > 0 else 0
        filas.append({
            'empleado': g['empleado'],
            'renta_acumulada': g['renta_acumulada'],
            'impuesto_acumulado': impuesto_acum,
            'meses': meses,
            'proyeccion_anual': proyeccion,
            'detalle': detalles[eid_emp],
        })

    # Sort by renta desc
    filas.sort(key=lambda x: x['renta_acumulada'], reverse=True)

    totales = {
        'renta_acumulada': sum(f['renta_acumulada'] for f in filas),
        'impuesto_acumulado': sum(f['impuesto_acumulado'] for f in filas),
        'proyeccion_anual': sum(f['proyeccion_anual'] for f in filas),
    }

    anios_disponibles = list(range(hoy.year - 2, hoy.year + 3))

    return render_template('remuneraciones/informe_renta.html',
        empresa=empresa, anio=anio, anios=anios_disponibles,
        filas=filas, totales=totales)


@bp.route('/empresa/<int:eid>/remuneraciones/buscar-liquidacion')
def buscar_liquidacion(eid):
    """Devuelve liquidaciones cuyo líquido coincide con el monto bancario (±1 peso)."""
    monto  = request.args.get('monto', 0, type=float)
    periodo = request.args.get('periodo', '').strip()
    q = (Liquidacion.query
         .join(Empleado, Liquidacion.empleado_id == Empleado.id)
         .filter(Liquidacion.empresa_id == eid)
         .filter(db.func.abs(Liquidacion.liquido - monto) < 1))
    if periodo:
        q = q.filter(Liquidacion.periodo == periodo)
    liqs = q.order_by(Liquidacion.periodo.desc()).limit(10).all()
    return jsonify([{
        'id':      l.id,
        'nombre':  l.empleado.nombre_completo,
        'periodo': l.periodo,
        'liquido': l.liquido,
        'url':     url_for('remuneraciones.detalle', eid=eid, liq_id=l.id),
    } for l in liqs])


@bp.route('/empresa/<int:eid>/remuneraciones/libro')
def libro(eid):
    from datetime import date
    empresa = Empresa.query.get_or_404(eid)
    hoy = date.today()
    periodo = request.args.get('periodo', f'{hoy.year}-{hoy.month:02d}')
    liqs = (Liquidacion.query
            .join(Empleado, Liquidacion.empleado_id == Empleado.id)
            .filter(Liquidacion.empresa_id == eid, Liquidacion.periodo == periodo)
            .order_by(Empleado.nombre)
            .all())
    periodos = [p[0] for p in (db.session.query(Liquidacion.periodo)
                .filter(Liquidacion.empresa_id == eid)
                .distinct().order_by(Liquidacion.periodo.desc()).all())]
    return render_template('remuneraciones/libro.html',
        empresa=empresa, periodo=periodo, liqs=liqs, periodos=periodos)


@bp.route('/empresa/<int:eid>/remuneraciones/libro/imprimir')
def libro_imprimir(eid):
    from datetime import date
    empresa = Empresa.query.get_or_404(eid)
    hoy = date.today()
    periodo = request.args.get('periodo', f'{hoy.year}-{hoy.month:02d}')
    liqs = (Liquidacion.query
            .join(Empleado, Liquidacion.empleado_id == Empleado.id)
            .filter(Liquidacion.empresa_id == eid, Liquidacion.periodo == periodo)
            .order_by(Empleado.nombre)
            .all())
    return render_template('remuneraciones/libro_imprimir.html',
        empresa=empresa, periodo=periodo, liqs=liqs, hoy=hoy)


@bp.route('/empresa/<int:eid>/remuneraciones/libro/excel')
def libro_excel(eid):
    from datetime import date
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    empresa = Empresa.query.get_or_404(eid)
    hoy = date.today()
    periodo = request.args.get('periodo', f'{hoy.year}-{hoy.month:02d}')
    liqs = (Liquidacion.query
            .join(Empleado, Liquidacion.empleado_id == Empleado.id)
            .filter(Liquidacion.empresa_id == eid, Liquidacion.periodo == periodo)
            .order_by(Empleado.nombre)
            .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Libro {periodo}'

    # Styles
    hdr_fill  = PatternFill('solid', fgColor='1F2937')
    hdr_font  = Font(color='FFFFFF', bold=True, size=9)
    tot_fill  = PatternFill('solid', fgColor='374151')
    tot_font  = Font(color='FFFFFF', bold=True, size=9)
    sub_fill  = PatternFill('solid', fgColor='E5E7EB')
    sub_font  = Font(bold=True, size=9)
    thin      = Side(style='thin', color='D1D5DB')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt   = '#,##0'
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right     = Alignment(horizontal='right', vertical='center')
    left_al   = Alignment(horizontal='left', vertical='center')

    # Title rows
    ws.merge_cells('A1:W1')
    ws['A1'] = f'LIBRO DE REMUNERACIONES – {empresa.razon_social} – RUT {empresa.rut}'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:W2')
    ws['A2'] = f'Período: {periodo}'
    ws['A2'].font = Font(size=10)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14

    # Column headers (row 3 = group labels, row 4 = individual labels)
    groups = [
        ('Identificación', 5),
        ('Haberes Imponibles', 4),
        ('No Imponibles', 2),
        ('Totales', 2),
        ('Descuentos Legales', 4),
        ('Total Desc. / Líquido', 2),
        ('Aportes Empleador', 4),
    ]
    col = 1
    for label, span in groups:
        start = get_column_letter(col)
        end   = get_column_letter(col + span - 1)
        if span > 1:
            ws.merge_cells(f'{start}3:{end}3')
        cell = ws[f'{start}3']
        cell.value = label
        cell.font  = sub_font
        cell.fill  = sub_fill
        cell.alignment = center
        cell.border = border
        col += span

    # Individual column headers (row 4)
    cols = [
        ('N°',         8,  center),
        ('RUT',        14, center),
        ('Nombre',     28, left_al),
        ('AFP',        10, center),
        ('Salud',      10, center),
        # Imponibles
        ('Sueldo\nBase', 14, right),
        ('HH.EE.',     11, right),
        ('Gratif.',    11, right),
        ('Otros\nImp.',11, right),
        # No imponibles
        ('Colación',   11, right),
        ('Moviliz.',   11, right),
        # Totales
        ('Total\nHaberes', 14, right),
        ('Renta\nImp.', 14, right),
        # Descuentos
        ('AFP $',      12, right),
        ('Salud $',    12, right),
        ('Ces.\nTrab.',11, right),
        ('Imp.\n2aCat',11, right),
        # Neto
        ('Total\nDesc.',12, right),
        ('Líquido',    14, right),
        # Empleador
        ('SIS',        11, right),
        ('Ces.\nEmp.', 11, right),
        ('Mutual',     11, right),
        ('Costo\nEmp.',14, right),
    ]
    for i, (label, width, align) in enumerate(cols, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = align; c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 30

    # Data rows
    for idx, liq in enumerate(liqs, 1):
        row = 4 + idx
        emp = liq.empleado
        salud_txt = emp.tipo_salud + (f' – {emp.isapre}' if emp.isapre else '')
        values = [
            idx, emp.rut, emp.nombre_completo, emp.afp, salud_txt,
            liq.sueldo_base, liq.horas_extra, liq.gratificacion, liq.otros_haberes,
            liq.bono_colacion, liq.bono_movilizacion,
            liq.total_haberes, liq.renta_imponible,
            liq.afp, liq.salud, liq.cesantia_trab, liq.impuesto_renta,
            liq.total_descuentos, liq.liquido,
            liq.sis, liq.cesantia_emp, liq.mutual, liq.costo_empresa,
        ]
        fill_row = PatternFill('solid', fgColor='F9FAFB') if idx % 2 == 0 else None
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(size=9)
            c.border = border
            if ci >= 6:
                c.number_format = num_fmt
                c.alignment = right
            else:
                c.alignment = left_al if ci == 3 else center
            if fill_row:
                c.fill = fill_row
        ws.row_dimensions[row].height = 14

    # Totals row
    if liqs:
        tot_row = 4 + len(liqs) + 1
        ws.cell(row=tot_row, column=1, value='TOTAL').font = tot_font
        ws.cell(row=tot_row, column=1).fill = tot_fill
        ws.cell(row=tot_row, column=1).alignment = center
        ws.cell(row=tot_row, column=1).border = border
        for ci in range(2, 6):
            c = ws.cell(row=tot_row, column=ci, value='')
            c.fill = tot_fill; c.border = border
        for ci in range(6, 24):
            col_letter = get_column_letter(ci)
            data_start  = 5
            data_end    = 4 + len(liqs)
            formula     = f'=SUM({col_letter}{data_start}:{col_letter}{data_end})'
            c = ws.cell(row=tot_row, column=ci, value=formula)
            c.font = tot_font; c.fill = tot_fill
            c.number_format = num_fmt; c.alignment = right; c.border = border
        ws.row_dimensions[tot_row].height = 16

    ws.freeze_panes = 'F5'

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'libro_remuneraciones_{empresa.rut}_{periodo}.xlsx'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ── Variables Mensuales (globales – sin empresa) ─────────────────────────────

@bp.route('/remuneraciones/variables')
def variables():
    next_url = request.args.get('next', '').strip()
    vars_list = VariablesMensuales.query.order_by(VariablesMensuales.periodo.desc()).all()
    return render_template('remuneraciones/variables.html', vars_list=vars_list, next_url=next_url)


@bp.route('/empresa/<int:eid>/remuneraciones/variables')
def variables_eid(eid):
    empresa = Empresa.query.get_or_404(eid)
    next_url = request.args.get('next', '').strip()
    vars_list = VariablesMensuales.query.order_by(VariablesMensuales.periodo.desc()).all()
    return render_template('remuneraciones/variables.html', vars_list=vars_list, empresa=empresa, next_url=next_url)


@bp.route('/remuneraciones/variables/guardar', methods=['POST'])
def variables_guardar():
    periodo = request.form.get('periodo', '').strip()
    if not periodo:
        flash('Período obligatorio', 'danger')
        return redirect(url_for('remuneraciones.variables'))

    v = VariablesMensuales.query.filter_by(periodo=periodo).first()
    if not v:
        v = VariablesMensuales(periodo=periodo)
        db.session.add(v)

    import json as _json
    def _f(name): return float(request.form.get(name, 0) or 0) or None
    v.uf = _f('uf')
    v.utm = _f('utm')
    v.tope_imponible = _f('tope_imponible')
    v.tope_gratificacion = _f('tope_gratificacion')
    v.imm = _f('imm')
    raw_sis = request.form.get('tasa_sis', '').strip()
    v.tasa_sis = float(raw_sis) / 100 if raw_sis else None

    # AFP commissions (submitted as afp_<Name> in %)
    AFP_NAMES = ['Capital', 'Cuprum', 'Habitat', 'Modelo', 'PlanVital', 'ProVida', 'Uno']
    tasas = {}
    for nombre in AFP_NAMES:
        raw = request.form.get(f'afp_{nombre}', '').strip()
        if raw:
            try:
                tasas[nombre] = round(float(raw.replace(',', '.')), 2)
            except ValueError:
                pass
    v.tasas_afp_json = _json.dumps(tasas) if tasas else None

    from datetime import datetime
    v.fecha_actualizacion = datetime.now()
    db.session.commit()
    flash(f'Variables {periodo} guardadas.', 'success')
    next_url = request.form.get('next', '').strip()
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect(url_for('remuneraciones.variables'))


@bp.route('/remuneraciones/variables/eliminar/<periodo>', methods=['POST'])
def variables_eliminar(periodo):
    v = VariablesMensuales.query.filter_by(periodo=periodo).first_or_404()
    db.session.delete(v)
    db.session.commit()
    flash(f'Variables {periodo} eliminadas.', 'success')
    return redirect(url_for('remuneraciones.variables'))


@bp.route('/empresa/<int:eid>/remuneraciones/variables/get/<periodo>')
def variables_get(eid, periodo):
    """Devuelve las variables de un período como JSON (para AJAX en liquidar)."""
    v = VariablesMensuales.query.filter_by(periodo=periodo).first()
    if not v:
        return jsonify({'ok': False})
    import json as _json
    tasas_afp = {}
    if v.tasas_afp_json:
        try:
            tasas_afp = _json.loads(v.tasas_afp_json)
        except Exception:
            pass
    return jsonify({
        'ok': True,
        'periodo': v.periodo,
        'uf': v.uf,
        'utm': v.utm,
        'tope_imponible': v.tope_imponible,
        'tope_gratificacion': v.tope_gratificacion,
        'imm': v.imm,
        'tasa_sis': round(v.tasa_sis * 100, 4) if v.tasa_sis else None,
        'tasas_afp': tasas_afp,
    })


@bp.route('/remuneraciones/variables/auto-fetch/<periodo>')
def variables_auto_fetch(periodo):
    """Retorna variables del período (DB o Previred). Si no están en DB, las obtiene y guarda."""
    import json as _json
    from datetime import datetime

    v = VariablesMensuales.query.filter_by(periodo=periodo).first()
    if v:
        return jsonify({
            'ok': True, 'source': 'db',
            'periodo': v.periodo, 'uf': v.uf, 'utm': v.utm,
            'tope_imponible': v.tope_imponible,
            'tope_gratificacion': v.tope_gratificacion,
        })

    from importers.previred import scrape
    try:
        data = scrape(periodo)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

    if not data.get('ok'):
        return jsonify(data)

    v = VariablesMensuales(periodo=periodo)
    db.session.add(v)
    v.uf = data.get('uf')
    v.utm = data.get('utm')
    v.tope_imponible = data.get('tope_imponible')
    v.tope_gratificacion = data.get('tope_gratificacion')
    v.imm = data.get('imm')
    v.tasa_sis = data.get('tasa_sis')
    tasas = data.get('tasas_afp') or {}
    v.tasas_afp_json = _json.dumps(tasas) if tasas else None
    v.fecha_actualizacion = datetime.now()
    db.session.commit()

    return jsonify({
        'ok': True, 'source': 'previred',
        'periodo': v.periodo, 'uf': v.uf, 'utm': v.utm,
        'tope_imponible': v.tope_imponible,
        'tope_gratificacion': v.tope_gratificacion,
    })


@bp.route('/remuneraciones/variables/fetch-indicadores')
def variables_fetch_previred():
    """Obtiene todos los indicadores previsionales desde previred.com."""
    from datetime import date
    from importers.previred import scrape
    try:
        periodo = request.args.get('periodo', '')
        if not periodo:
            periodo = date.today().strftime('%Y-%m')
        return jsonify(scrape(periodo))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ── Tabla UF diaria ──────────────────────────────────────────────────────────

@bp.route('/remuneraciones/uf-tabla')
def uf_tabla():
    from datetime import date
    import calendar
    hoy = date.today()
    anio = request.args.get('anio', hoy.year, type=int)
    desde = date(anio, 1, 1)
    hasta = date(anio, 12, 31)
    filas = (ValorUF.query
             .filter(ValorUF.fecha >= desde, ValorUF.fecha <= hasta)
             .order_by(ValorUF.fecha)
             .all())
    # dict date -> valor for fast lookup in template
    uf_dict = {f.fecha: f.valor for f in filas}
    # expected days = all days up to today (or end of year if past)
    limite = min(hasta, hoy)
    n_esperados = (limite - desde).days + 1 if limite >= desde else 0
    n_cargados = len(filas)
    # per-month info: list of (mes_num, nombre, [(day, valor|None), ...])
    MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    meses_data = []
    for m in range(1, 13):
        ultimo_dia = calendar.monthrange(anio, m)[1]
        dias = []
        for d in range(1, ultimo_dia + 1):
            fd = date(anio, m, d)
            dias.append((d, uf_dict.get(fd), fd))
        meses_data.append((m, MESES[m-1], dias))
    return render_template('remuneraciones/uf_tabla.html',
                           meses_data=meses_data, uf_dict=uf_dict,
                           anio=anio, hoy=hoy,
                           n_esperados=n_esperados, n_cargados=n_cargados)


@bp.route('/remuneraciones/uf-tabla/actualizar', methods=['POST'])
def uf_actualizar():
    """Descarga todos los valores UF del año desde mindicador.cl."""
    from datetime import date as dt
    from services.uf import fetch_year
    hoy = dt.today()
    anio = request.form.get('anio', hoy.year, type=int)
    try:
        actualizados = fetch_year(anio)
        flash(f'{actualizados} valores UF actualizados para {anio}.', 'success')
    except Exception as e:
        flash(f'Error al conectar con mindicador.cl: {e}', 'danger')
    return redirect(url_for('remuneraciones.uf_tabla', anio=anio))


@bp.route('/remuneraciones/uf-tabla/guardar', methods=['POST'])
def uf_guardar():
    """Guarda o actualiza un valor UF manual."""
    from datetime import date as dt
    raw = request.form.get('fecha', '').strip()
    valor_s = request.form.get('valor', '').strip().replace(',', '.')
    try:
        fecha = dt.fromisoformat(raw)
        valor = float(valor_s)
    except (ValueError, TypeError):
        flash('Fecha o valor inválido.', 'danger')
        return redirect(url_for('remuneraciones.uf_tabla'))
    existing = ValorUF.query.filter_by(fecha=fecha).first()
    if existing:
        existing.valor = valor
    else:
        db.session.add(ValorUF(fecha=fecha, valor=valor))
    db.session.commit()
    flash(f'UF {fecha} = ${valor:,.2f} guardada.', 'success')
    return redirect(url_for('remuneraciones.uf_tabla', anio=fecha.year, mes=fecha.month))


@bp.route('/remuneraciones/uf/get')
def uf_get_valor():
    """Retorna el valor UF para una fecha (JSON). Usado por otras secciones."""
    from datetime import date as dt
    raw = request.args.get('fecha', '')
    try:
        fecha = dt.fromisoformat(raw)
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'fecha inválida'})
    # Busca la fecha exacta; si no existe, retorna la más reciente anterior
    uf = (ValorUF.query
          .filter(ValorUF.fecha <= fecha)
          .order_by(ValorUF.fecha.desc())
          .first())
    if uf:
        return jsonify({'ok': True, 'fecha': str(uf.fecha), 'valor': uf.valor})
    return jsonify({'ok': False, 'error': 'Sin datos UF para esa fecha'})


def _poblar(emp, form):
    emp.rut = form.get('rut', '').strip()
    emp.nombre = form.get('nombre', '').strip()
    emp.apellido_paterno = form.get('apellido_paterno', '').strip() or None
    emp.apellido_materno = form.get('apellido_materno', '').strip() or None
    emp.cargo = form.get('cargo', '').strip()
    emp.tipo_contrato = form.get('tipo_contrato', 'INDEFINIDO')
    raw_sueldo = (form.get('sueldo_base', '0') or '0').replace('.', '').replace(',', '.')
    emp.sueldo_base = float(raw_sueldo) if raw_sueldo else 0.0
    emp.tipo_sueldo = form.get('tipo_sueldo', 'BRUTO')
    emp.afp = form.get('afp', 'Habitat')
    raw_tasa = form.get('tasa_afp_comision', '').strip()
    emp.tasa_afp_comision = float(raw_tasa) / 100 if raw_tasa else motor.AFP_COMISIONES.get(emp.afp, 0.0127)
    emp.tipo_salud = form.get('tipo_salud', 'FONASA')
    emp.isapre = form.get('isapre', '').strip() or None
    raw_isapre_uf = form.get('monto_isapre_uf', '').strip()
    emp.monto_isapre_uf = float(raw_isapre_uf) if raw_isapre_uf else 0.0
    # Keep legacy monto_isapre in sync (set to 0 since we use UF now)
    emp.monto_isapre = 0.0
    raw_colacion = (form.get('bono_colacion', '0') or '0').replace('.', '').replace(',', '.')
    emp.bono_colacion = float(raw_colacion) if raw_colacion else 0.0
    raw_movil = (form.get('bono_movilizacion', '0') or '0').replace('.', '').replace(',', '.')
    emp.bono_movilizacion = float(raw_movil) if raw_movil else 0.0
    raw_otros = (form.get('otros_haberes', '0') or '0').replace('.', '').replace(',', '.')
    emp.otros_haberes = float(raw_otros) if raw_otros else 0.0
    raw_mutual = form.get('tasa_mutual', '').strip()
    emp.tasa_mutual = float(raw_mutual) / 100 if raw_mutual else 0.0093
    raw_apv = (form.get('apv_monto', '0') or '0').replace('.', '').replace(',', '.')
    emp.apv_monto = float(raw_apv) if raw_apv else 0.0
    emp.apv_tipo = form.get('apv_tipo', 'A') or 'A'
    emp.activo = form.get('activo') == 'on'
    fi = form.get('fecha_ingreso', '').strip()
    if fi:
        from datetime import date
        try:
            emp.fecha_ingreso = date.fromisoformat(fi)
        except ValueError:
            pass


# ── Documentos de empleado (contratos, anexos, etc.) ────────────────────────

@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/documentos')
def docs_empleado(eid, emp_id):
    from models import DocumentoEmpleado
    empresa = Empresa.query.get_or_404(eid)
    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()
    docs = (DocumentoEmpleado.query
            .filter_by(empleado_id=emp_id)
            .order_by(DocumentoEmpleado.fecha_documento.desc(),
                      DocumentoEmpleado.creado_en.desc())
            .all())
    return render_template('remuneraciones/docs_empleado.html',
                           empresa=empresa, emp=emp, docs=docs)


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/documentos/subir', methods=['POST'])
def docs_empleado_subir(eid, emp_id):
    from models import DocumentoEmpleado
    from storage import save_attachment
    import re

    emp = Empleado.query.filter_by(id=emp_id, empresa_id=eid).first_or_404()
    empresa = Empresa.query.get_or_404(eid)
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash('Debes seleccionar un archivo.', 'danger')
        return redirect(url_for('remuneraciones.docs_empleado', eid=eid, emp_id=emp_id))

    rut_limpio = re.sub(r'[^\w]', '_', empresa.rut)
    subfolder = f"{rut_limpio}/contratos/emp_{emp_id}"
    try:
        storage_key = save_attachment(
            archivo, archivo.filename,
            current_app.config['UPLOAD_FOLDER'],
            subfolder=subfolder,
        )
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('remuneraciones.docs_empleado', eid=eid, emp_id=emp_id))

    from datetime import date as _date
    raw_fecha = request.form.get('fecha_documento', '').strip()
    try:
        fecha_doc = _date.fromisoformat(raw_fecha) if raw_fecha else None
    except ValueError:
        fecha_doc = None

    doc = DocumentoEmpleado(
        empleado_id=emp_id,
        empresa_id=eid,
        tipo=request.form.get('tipo', 'CONTRATO'),
        descripcion=request.form.get('descripcion', '').strip() or None,
        fecha_documento=fecha_doc,
        archivo_url=storage_key,
    )
    db.session.add(doc)
    db.session.commit()
    flash('Documento subido correctamente.', 'success')
    return redirect(url_for('remuneraciones.docs_empleado', eid=eid, emp_id=emp_id))


@bp.route('/empresa/<int:eid>/remuneraciones/<int:emp_id>/documentos/<int:doc_id>/eliminar', methods=['POST'])
def docs_empleado_eliminar(eid, doc_id, emp_id):
    from models import DocumentoEmpleado
    doc = DocumentoEmpleado.query.filter_by(id=doc_id, empresa_id=eid).first_or_404()
    db.session.delete(doc)
    db.session.commit()
    flash('Documento eliminado.', 'success')
    return redirect(url_for('remuneraciones.docs_empleado', eid=eid, emp_id=emp_id))

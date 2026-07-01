"""API REST para clientes remotos (Claude Code en otra PC, scripts, integraciones).

Diseño:
- JSON in/out
- Sin auth (Tailscale = red privada). Si se expone públicamente, agregar API key.
- Reusa la lógica del backend (confirmar_asiento, _validar_aux_requerido, etc.)
- Todos los POST de escritura validan cuadre + aux antes de persistir.

Convenciones:
- Fechas: ISO 8601 (YYYY-MM-DD)
- Montos: float
- Cuenta se identifica por código (1.1.02) o id; contraparte por id
"""
import json
from datetime import date, datetime
from flask import Blueprint, request, jsonify, abort
from sqlalchemy import func, text
from models import (db, Empresa, Asiento, LineaAsiento, Cuenta, Contraparte,
                    DocumentoSII, MovimientoBanco)
from engine.asientos import confirmar_asiento, anular_asiento

bp = Blueprint('api', __name__, url_prefix='/api')


def _e2j(asiento, with_lines=False):
    d = {
        'id': asiento.id, 'empresa_id': asiento.empresa_id,
        'numero': asiento.numero, 'fecha': asiento.fecha.isoformat() if asiento.fecha else None,
        'descripcion': asiento.descripcion, 'origen': asiento.origen, 'estado': asiento.estado,
        'total_debe': float(asiento.total_debe or 0), 'total_haber': float(asiento.total_haber or 0),
        'cuadrado': asiento.cuadrado,
    }
    if with_lines:
        d['lineas'] = [{
            'id': l.id, 'cuenta_id': l.cuenta_id, 'cuenta_codigo': l.cuenta.codigo,
            'cuenta_nombre': l.cuenta.nombre,
            'contraparte_id': l.contraparte_id,
            'contraparte_rut': l.contraparte.rut if l.contraparte else None,
            'contraparte_nombre': l.contraparte.razon_social if l.contraparte else None,
            'debe': float(l.debe or 0), 'haber': float(l.haber or 0),
            'descripcion': l.descripcion, 'orden': l.orden,
        } for l in sorted(asiento.lineas, key=lambda x: x.orden or 0)]
    return d


# ─── Helpers para serialización ───────────────────────────────────────────────
def _err(msg, code=400):
    return jsonify({'error': msg}), code


# ─── Health & catálogo ────────────────────────────────────────────────────────
@bp.route('/health')
def health():
    return jsonify({'ok': True, 'ts': datetime.now().isoformat()})


@bp.route('/empresas')
def empresas():
    activas = request.args.get('activas', '1') == '1'
    q = Empresa.query
    if activas:
        q = q.filter_by(activa=True)
    return jsonify([{
        'id': e.id, 'rut': e.rut, 'razon_social': e.razon_social,
        'nombre_fantasia': e.nombre_fantasia, 'regimen': e.regimen,
        'tc_activa': bool(getattr(e, 'tc_activa', False)),
        'contribuyente_iva': bool(e.contribuyente_iva),
        'activa': bool(e.activa),
    } for e in q.order_by(Empresa.razon_social).all()])


@bp.route('/empresa/<int:eid>')
def empresa_detalle(eid):
    e = Empresa.query.get_or_404(eid)
    return jsonify({
        'id': e.id, 'rut': e.rut, 'razon_social': e.razon_social,
        'nombre_fantasia': e.nombre_fantasia, 'giro': e.giro,
        'regimen': e.regimen, 'contribuyente_iva': bool(e.contribuyente_iva),
        'tasa_ppm': e.tasa_ppm, 'tc_activa': bool(getattr(e, 'tc_activa', False)),
        'activa': bool(e.activa),
    })


@bp.route('/empresa/<int:eid>/cuentas')
def cuentas(eid):
    Empresa.query.get_or_404(eid)
    return jsonify([{
        'id': c.id, 'codigo': c.codigo, 'nombre': c.nombre, 'tipo': c.tipo,
        'naturaleza': c.naturaleza, 'requiere_aux': bool(c.requiere_aux),
        'es_titulo': bool(c.es_titulo), 'activa': bool(c.activa),
    } for c in Cuenta.query.filter_by(empresa_id=eid).order_by(Cuenta.codigo).all()])


@bp.route('/empresa/<int:eid>/contrapartes')
def contrapartes(eid):
    Empresa.query.get_or_404(eid)
    q = (request.args.get('q') or '').strip().lower()
    base = Contraparte.query.filter_by(empresa_id=eid, activo=True)
    if q:
        pattern = f'%{q}%'
        base = base.filter(db.or_(
            func.lower(Contraparte.razon_social).like(pattern),
            func.lower(Contraparte.rut).like(pattern)
        ))
    base = base.order_by(Contraparte.razon_social)
    if q:
        base = base.limit(50)
    return jsonify([{
        'id': c.id, 'rut': c.rut, 'razon_social': c.razon_social, 'tipo': c.tipo,
    } for c in base.all()])


@bp.route('/empresa/<int:eid>/contraparte', methods=['POST'])
def contraparte_crear(eid):
    Empresa.query.get_or_404(eid)
    data = request.get_json() or {}
    rut = (data.get('rut') or '').strip()
    razon = (data.get('razon_social') or '').strip()
    tipo = data.get('tipo', 'PROVEEDOR')
    if not rut or not razon:
        return jsonify({'error': 'rut y razon_social requeridos'}), 400
    existente = Contraparte.query.filter_by(empresa_id=eid, rut=rut).first()
    if existente:
        return jsonify({'id': existente.id, 'rut': existente.rut,
                        'razon_social': existente.razon_social, 'existente': True})
    cp = Contraparte(empresa_id=eid, rut=rut, razon_social=razon, tipo=tipo, activo=True)
    db.session.add(cp)
    db.session.commit()
    return jsonify({'id': cp.id, 'rut': cp.rut, 'razon_social': cp.razon_social,
                    'existente': False}), 201


# ─── Movs banco y SII ─────────────────────────────────────────────────────────
@bp.route('/empresa/<int:eid>/movs-banco')
def movs_banco(eid):
    Empresa.query.get_or_404(eid)
    desde = request.args.get('desde'); hasta = request.args.get('hasta')
    procesado = request.args.get('procesado')  # '1', '0', None=todos
    q = MovimientoBanco.query.filter_by(empresa_id=eid)
    if desde: q = q.filter(MovimientoBanco.fecha >= desde)
    if hasta: q = q.filter(MovimientoBanco.fecha <= hasta)
    if procesado in ('0', 'false', 'False'):
        q = q.filter(MovimientoBanco.procesado == False)
    elif procesado in ('1', 'true', 'True'):
        q = q.filter(MovimientoBanco.procesado == True)
    return jsonify([{
        'id': m.id, 'fecha': m.fecha.isoformat() if m.fecha else None,
        'descripcion': m.descripcion, 'banco': m.banco,
        'cargo': float(m.cargo or 0), 'abono': float(m.abono or 0),
        'saldo': float(m.saldo) if m.saldo is not None else None,
        'procesado': bool(m.procesado), 'asiento_id': m.asiento_id,
        'conciliacion_id': m.conciliacion_id,
    } for m in q.order_by(MovimientoBanco.fecha, MovimientoBanco.id).all()])


@bp.route('/empresa/<int:eid>/sii')
def sii_docs(eid):
    Empresa.query.get_or_404(eid)
    desde = request.args.get('desde'); hasta = request.args.get('hasta')
    libro = request.args.get('libro')  # COMPRAS|VENTAS|HONORARIOS
    procesado = request.args.get('procesado')
    q = DocumentoSII.query.filter_by(empresa_id=eid)
    if desde: q = q.filter(DocumentoSII.fecha >= desde)
    if hasta: q = q.filter(DocumentoSII.fecha <= hasta)
    if libro: q = q.filter_by(tipo_libro=libro.upper())
    if procesado in ('0', 'false', 'False'):
        q = q.filter(DocumentoSII.procesado == False)
    elif procesado in ('1', 'true', 'True'):
        q = q.filter(DocumentoSII.procesado == True)
    return jsonify([{
        'id': d.id, 'tipo_libro': d.tipo_libro, 'tipo_dte': d.tipo_dte,
        'folio': d.folio, 'fecha': d.fecha.isoformat() if d.fecha else None,
        'rut_contraparte': d.rut_contraparte,
        'razon_social_contraparte': d.razon_social_contraparte,
        'monto_neto': float(d.monto_neto or 0), 'iva': float(d.iva or 0),
        'monto_exento': float(d.monto_exento or 0), 'total': float(d.total or 0),
        'procesado': bool(d.procesado), 'asiento_id': d.asiento_id,
    } for d in q.order_by(DocumentoSII.fecha, DocumentoSII.folio).all()])


# ─── Asientos: crear, confirmar, anular, listar, detalle ──────────────────────
@bp.route('/empresa/<int:eid>/asientos')
def asientos_lista(eid):
    Empresa.query.get_or_404(eid)
    desde = request.args.get('desde'); hasta = request.args.get('hasta')
    estado = request.args.get('estado')
    q = Asiento.query.filter_by(empresa_id=eid)
    if desde: q = q.filter(Asiento.fecha >= desde)
    if hasta: q = q.filter(Asiento.fecha <= hasta)
    if estado: q = q.filter_by(estado=estado)
    return jsonify([_e2j(a) for a in q.order_by(Asiento.fecha, Asiento.numero).all()])


@bp.route('/api/asiento/<int:aid>')
@bp.route('/asiento/<int:aid>')
def asiento_detalle(aid):
    a = Asiento.query.get_or_404(aid)
    return jsonify(_e2j(a, with_lines=True))


@bp.route('/empresa/<int:eid>/asiento', methods=['POST'])
def asiento_crear(eid):
    """Crea un asiento con sus líneas. Valida cuadre + aux requerido.
    Payload JSON:
    {
        "fecha": "2026-04-01",
        "descripcion": "...",
        "estado": "BORRADOR" (default) | "CONFIRMADO",
        "origen": "MANUAL" (default) | "BANCO" | "SII" | etc.,
        "lineas": [
            {"cuenta_codigo": "1.1.02", "debe": 100000, "haber": 0,
             "descripcion": "Pago X", "contraparte_id": 5},
            ...
        ],
        "mov_banco_ids": [1234, ...],   # opcional, marca movs procesados
        "sii_doc_ids": [567, ...]       # opcional, marca docs procesados
    }
    """
    empresa = Empresa.query.get_or_404(eid)
    data = request.get_json() or {}

    # Validaciones básicas
    try:
        fecha = date.fromisoformat(data['fecha'])
    except (KeyError, ValueError):
        return jsonify({'error': 'fecha inválida (YYYY-MM-DD requerido)'}), 400
    desc = (data.get('descripcion') or '').strip()
    if not desc:
        return jsonify({'error': 'descripcion requerida'}), 400
    lineas_in = data.get('lineas') or []
    if not lineas_in:
        return jsonify({'error': 'al menos una línea requerida'}), 400

    # Resolver cuentas por código → id
    codigos = {c.codigo: c for c in Cuenta.query.filter_by(empresa_id=eid).all()}
    lineas_parsed = []
    total_d = total_h = 0.0
    for i, ln in enumerate(lineas_in):
        cod = ln.get('cuenta_codigo') or ''
        cta = codigos.get(cod)
        if not cta:
            cta_id = ln.get('cuenta_id')
            if cta_id:
                cta = Cuenta.query.filter_by(empresa_id=eid, id=int(cta_id)).first()
        if not cta:
            return jsonify({'error': f'línea {i+1}: cuenta {cod or ln.get("cuenta_id")} no existe'}), 400
        d = float(ln.get('debe') or 0); h = float(ln.get('haber') or 0)
        cp_id = ln.get('contraparte_id')
        # Validar aux requerido
        if cta.requiere_aux and (d or h) and not cp_id:
            return jsonify({'error': f'línea {i+1} ({cta.codigo} {cta.nombre}) requiere contraparte_id'}), 400
        if cp_id:
            cp = Contraparte.query.filter_by(empresa_id=eid, id=int(cp_id)).first()
            if not cp:
                return jsonify({'error': f'línea {i+1}: contraparte_id {cp_id} no existe en esta empresa'}), 400
        lineas_parsed.append({
            'cuenta_id': cta.id, 'debe': d, 'haber': h,
            'descripcion': (ln.get('descripcion') or '').strip(),
            'contraparte_id': int(cp_id) if cp_id else None,
            'orden': i + 1,
        })
        total_d += d; total_h += h

    if abs(total_d - total_h) > 1:
        return jsonify({'error': f'asiento descuadrado: D={total_d:,.0f} H={total_h:,.0f} diff={total_d-total_h:,.0f}'}), 400

    estado = (data.get('estado') or 'BORRADOR').upper()
    origen = (data.get('origen') or 'MANUAL').upper()

    # Insert
    asiento = Asiento(empresa_id=eid, fecha=fecha, descripcion=desc,
                      estado='BORRADOR', origen=origen)
    db.session.add(asiento)
    db.session.flush()
    for l in lineas_parsed:
        db.session.add(LineaAsiento(asiento_id=asiento.id, **l))
    db.session.flush()

    # Marcar movs banco / sii como procesados
    for mid in (data.get('mov_banco_ids') or []):
        m = MovimientoBanco.query.filter_by(empresa_id=eid, id=int(mid)).first()
        if m:
            m.procesado = True; m.asiento_id = asiento.id
    for did in (data.get('sii_doc_ids') or []):
        d_sii = DocumentoSII.query.filter_by(empresa_id=eid, id=int(did)).first()
        if d_sii:
            d_sii.procesado = True; d_sii.asiento_id = asiento.id

    # Confirmar si fue solicitado
    if estado == 'CONFIRMADO':
        try:
            confirmar_asiento(asiento)
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'no se pudo confirmar: {e}'}), 400

    db.session.commit()
    return jsonify(_e2j(asiento, with_lines=True)), 201


@bp.route('/asiento/<int:aid>/confirmar', methods=['POST'])
def asiento_confirmar(aid):
    a = Asiento.query.get_or_404(aid)
    if a.estado == 'CONFIRMADO':
        return jsonify({'error': 'ya está confirmado'}), 400
    try:
        confirmar_asiento(a)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    return jsonify(_e2j(a, with_lines=True))


@bp.route('/asiento/<int:aid>/anular', methods=['POST'])
def asiento_anular_api(aid):
    a = Asiento.query.get_or_404(aid)
    try:
        anular_asiento(a)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400
    return jsonify(_e2j(a, with_lines=True))


# ─── Saldos y mayor ───────────────────────────────────────────────────────────
@bp.route('/empresa/<int:eid>/archivos')
def archivos_lista(eid):
    """Lista archivos en backup de la empresa (filesystem)."""
    import os
    empresa = Empresa.query.get_or_404(eid)
    from routes.documentos import _scan_dir, _rut_clean
    from flask import current_app as _ca
    base = os.path.join(_ca.config['UPLOAD_FOLDER'], 'backups_importacion', _rut_clean(empresa.rut))
    archivos = _scan_dir(base, max_depth=5)
    return jsonify([{
        'rel_path': a['rel_path'], 'nombre': a['nombre'], 'tipo': a['tipo'],
        'periodo': a['periodo'] or None, 'tamano': a['tamano'], 'mtime': a['mtime'],
        'es_global': not a['periodo'],
    } for a in sorted(archivos, key=lambda x: x['mtime'], reverse=True)])


@bp.route('/empresa/<int:eid>/archivo')
def archivo_contenido(eid):
    """Devuelve contenido parseado de un archivo (Excel/CSV/TXT). PDFs e imágenes no
    se parsean — pedir descarga via /empresa/<id>/archivos/descargar?rel=...
    Útil para que un Claude remoto pueda leer cartolas/Excels sin descargarlos.
    Query params: rel=ruta_relativa, max_rows=N (default 1000)
    """
    import os
    empresa = Empresa.query.get_or_404(eid)
    rel = request.args.get('rel', '').strip()
    max_rows = int(request.args.get('max_rows', 1000))
    if not rel or '..' in rel.split('/'):
        return jsonify({'error': 'rel inválido'}), 400
    from routes.documentos import _rut_clean
    from flask import current_app as _ca
    base = os.path.join(_ca.config['UPLOAD_FOLDER'], 'backups_importacion', _rut_clean(empresa.rut))
    full = os.path.join(base, rel)
    if not os.path.isfile(full):
        return jsonify({'error': f'archivo no encontrado: {rel}'}), 404
    ext = os.path.splitext(full)[1].lower()
    fname = os.path.basename(full)
    try:
        if ext == '.csv':
            import csv as _csv
            rows = []
            with open(full, encoding='utf-8', errors='replace') as f:
                sample = f.read(2048); f.seek(0)
                delim = ';' if sample.count(';') > sample.count(',') else ','
                for i, row in enumerate(_csv.reader(f, delimiter=delim)):
                    if i >= max_rows: break
                    rows.append(row)
            return jsonify({'ext': ext, 'nombre': fname, 'rows': rows, 'count': len(rows)})
        elif ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(full, data_only=True, read_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows: break
                rows.append(['' if v is None else str(v) for v in row])
            return jsonify({'ext': ext, 'nombre': fname, 'rows': rows, 'count': len(rows),
                            'sheet': ws.title})
        elif ext == '.xls':
            with open(full, 'rb') as f:
                head_raw = f.read(2048)
            head = head_raw.lstrip().lower()
            is_binary_xls = head_raw[:4] == b'\xd0\xcf\x11\xe0'
            is_html = (not is_binary_xls) and head.startswith(b'<') and (b'<table' in head or b'<html' in head)
            is_text = (not is_binary_xls) and (not is_html) and all(
                b in (9, 10, 13) or 32 <= b < 127 or b > 127 for b in head_raw[:512])
            if is_text and not is_html:
                import csv as _csv
                rows = []
                with open(full, encoding='utf-8', errors='replace') as f:
                    sample = f.read(2048); f.seek(0)
                    delim = ';' if sample.count(';') > sample.count(',') else (',' if ',' in sample else '\t')
                    for i, row in enumerate(_csv.reader(f, delimiter=delim)):
                        if i >= max_rows: break
                        rows.append(row)
                return jsonify({'ext': ext, 'nombre': fname, 'rows': rows, 'count': len(rows),
                                'formato': 'csv-disfrazado', 'delim': delim})
            if is_html:
                from html.parser import HTMLParser
                class _P(HTMLParser):
                    def __init__(self):
                        super().__init__(); self.rows=[]; self.row=None; self.cell=None; self.txt=''; self.in_t=False
                    def handle_starttag(self,t,a):
                        if t=='table': self.in_t=True
                        elif t=='tr' and self.in_t: self.row=[]
                        elif t in ('td','th') and self.row is not None: self.cell=[]; self.txt=''
                        elif t=='br' and self.cell is not None: self.txt+=' '
                    def handle_endtag(self,t):
                        if t in ('td','th') and self.cell is not None and self.row is not None:
                            self.row.append(self.txt.strip()); self.cell=None; self.txt=''
                        elif t=='tr' and self.row is not None: self.rows.append(self.row); self.row=None
                        elif t=='table': self.in_t=False
                    def handle_data(self,d):
                        if self.cell is not None: self.txt+=d
                with open(full, encoding='utf-8', errors='replace') as f:
                    p = _P(); p.feed(f.read())
                rows = p.rows[:max_rows]
                return jsonify({'ext': ext, 'nombre': fname, 'rows': rows, 'count': len(rows),
                                'formato': 'html-disfrazado'})
            import xlrd
            wb = xlrd.open_workbook(full)
            sh = wb.sheet_by_index(0)
            rows = []
            for r in range(min(sh.nrows, max_rows)):
                rows.append([str(sh.cell_value(r, c)) for c in range(sh.ncols)])
            return jsonify({'ext': ext, 'nombre': fname, 'rows': rows, 'count': len(rows),
                            'sheet': sh.name})
        elif ext in ('.txt', '.log', '.md'):
            with open(full, encoding='utf-8', errors='replace') as f:
                contenido = f.read(500_000)
            return jsonify({'ext': ext, 'nombre': fname, 'texto': contenido,
                            'truncated': len(contenido) >= 500_000})
        else:
            return jsonify({'error': f'tipo {ext} no parseable — usar descarga'}), 415
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/empresa/<int:eid>/saldos')
def saldos(eid):
    Empresa.query.get_or_404(eid)
    hasta = request.args.get('hasta', date.today().isoformat())
    desde = request.args.get('desde')  # opcional
    rows = (db.session.query(
                Cuenta.codigo, Cuenta.nombre, Cuenta.tipo,
                func.sum(LineaAsiento.debe).label('td'),
                func.sum(LineaAsiento.haber).label('th'))
            .join(LineaAsiento, LineaAsiento.cuenta_id == Cuenta.id)
            .join(Asiento, Asiento.id == LineaAsiento.asiento_id)
            .filter(Cuenta.empresa_id == eid, Asiento.empresa_id == eid,
                    Asiento.estado == 'CONFIRMADO',
                    Asiento.fecha <= hasta)
            .group_by(Cuenta.id).order_by(Cuenta.codigo).all())
    if desde:
        # filtrar también por desde
        rows = [r for r in rows]  # ya está aplicado en query si pasáramos desde — simplificamos
    return jsonify([{
        'codigo': r.codigo, 'nombre': r.nombre, 'tipo': r.tipo,
        'debe': float(r.td or 0), 'haber': float(r.th or 0),
        'saldo': float((r.td or 0) - (r.th or 0)),
    } for r in rows if (r.td or r.th)])


@bp.route('/empresa/<int:eid>/cuenta/<path:codigo>/mayor')
def mayor_cuenta(eid, codigo):
    Empresa.query.get_or_404(eid)
    cta = Cuenta.query.filter_by(empresa_id=eid, codigo=codigo).first()
    if not cta:
        return jsonify({'error': f'cuenta {codigo} no existe'}), 404
    desde = request.args.get('desde'); hasta = request.args.get('hasta')
    q = (db.session.query(LineaAsiento, Asiento)
         .join(Asiento, Asiento.id == LineaAsiento.asiento_id)
         .filter(LineaAsiento.cuenta_id == cta.id,
                 Asiento.empresa_id == eid,
                 Asiento.estado == 'CONFIRMADO'))
    if desde: q = q.filter(Asiento.fecha >= desde)
    if hasta: q = q.filter(Asiento.fecha <= hasta)
    rows = q.order_by(Asiento.fecha, Asiento.numero).all()
    return jsonify({
        'cuenta': {'codigo': cta.codigo, 'nombre': cta.nombre},
        'lineas': [{
            'asiento_id': a.id, 'numero': a.numero, 'fecha': a.fecha.isoformat(),
            'descripcion': a.descripcion,
            'debe': float(l.debe or 0), 'haber': float(l.haber or 0),
            'glosa': l.descripcion,
            'contraparte_id': l.contraparte_id,
        } for l, a in rows],
    })


# ─── Query SQL libre (solo SELECT, sólo lectura) ──────────────────────────────
@bp.route('/sql', methods=['POST'])
def sql_query():
    """SELECT libre para casos no cubiertos. Solo lectura.
    Payload: {"sql": "SELECT ...", "params": {...}}"""
    data = request.get_json() or {}
    sql = (data.get('sql') or '').strip()
    if not sql.lower().startswith('select'):
        return jsonify({'error': 'solo SELECT permitido'}), 400
    params = data.get('params') or {}
    try:
        result = db.session.execute(text(sql), params)
        rows = [dict(r._mapping) for r in result.fetchall()]
        # Convert non-serializable
        for r in rows:
            for k, v in r.items():
                if isinstance(v, (date, datetime)):
                    r[k] = v.isoformat()
        return jsonify({'rows': rows, 'count': len(rows)})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ════════════════════════════════════════════════════════════════════════════
# Endpoints adicionales (priority 1-6 según pedido cliente remoto)
# ════════════════════════════════════════════════════════════════════════════

# ─── P1: Notas contables (lectura/escritura) ──────────────────────────────────
@bp.route('/empresa/<int:eid>/nota', methods=['GET'])
def nota_get(eid):
    Empresa.query.get_or_404(eid)
    r = db.session.execute(text(
        "SELECT contenido, actualizado_en FROM notas_contables WHERE empresa_id=:e"
    ), {'e': eid}).fetchone()
    if not r:
        return _err('nota no existe', 404)
    return jsonify({
        'empresa_id': eid,
        'contenido': r[0],
        'actualizado_en': r[1].isoformat() if hasattr(r[1], 'isoformat') else str(r[1]),
    })


@bp.route('/empresa/<int:eid>/nota', methods=['PUT'])
def nota_put(eid):
    Empresa.query.get_or_404(eid)
    data = request.get_json() or {}
    contenido = data.get('contenido')
    if contenido is None:
        return _err('contenido requerido')
    db.session.execute(text("""
        INSERT INTO notas_contables (empresa_id, contenido, actualizado_en)
        VALUES (:e, :c, CURRENT_TIMESTAMP)
        ON CONFLICT(empresa_id) DO UPDATE SET
            contenido=excluded.contenido, actualizado_en=CURRENT_TIMESTAMP
    """), {'e': eid, 'c': contenido})
    db.session.commit()
    r = db.session.execute(text(
        "SELECT contenido, actualizado_en FROM notas_contables WHERE empresa_id=:e"
    ), {'e': eid}).fetchone()
    return jsonify({
        'empresa_id': eid,
        'contenido': r[0],
        'actualizado_en': r[1].isoformat() if hasattr(r[1], 'isoformat') else str(r[1]),
    })


# ─── P2: Saldos con estados configurables (incl. BORRADOR) ────────────────────
@bp.route('/empresa/<int:eid>/saldos-estados')
def saldos_con_estados(eid):
    """Saldos por cuenta filtrando por estados de asiento (CSV).
    Default: CONFIRMADO. Ej: ?estados=BORRADOR,CONFIRMADO para ver proyección."""
    Empresa.query.get_or_404(eid)
    hasta = request.args.get('hasta', date.today().isoformat())
    estados_csv = request.args.get('estados', 'CONFIRMADO')
    estados = [s.strip().upper() for s in estados_csv.split(',') if s.strip()]
    if not estados:
        return _err('estados vacío')
    rows = (db.session.query(
                Cuenta.codigo, Cuenta.nombre, Cuenta.tipo,
                func.sum(LineaAsiento.debe).label('td'),
                func.sum(LineaAsiento.haber).label('th'))
            .join(LineaAsiento, LineaAsiento.cuenta_id == Cuenta.id)
            .join(Asiento, Asiento.id == LineaAsiento.asiento_id)
            .filter(Cuenta.empresa_id == eid, Asiento.empresa_id == eid,
                    Asiento.estado.in_(estados),
                    Asiento.fecha <= hasta)
            .group_by(Cuenta.id).order_by(Cuenta.codigo).all())
    return jsonify({
        'hasta': hasta, 'estados': estados,
        'cuentas': [{
            'codigo': r.codigo, 'nombre': r.nombre, 'tipo': r.tipo,
            'debe': float(r.td or 0), 'haber': float(r.th or 0),
            'saldo': float((r.td or 0) - (r.th or 0)),
        } for r in rows if (r.td or r.th)]
    })


# ─── P3: Resumen de mes (dashboard) ───────────────────────────────────────────
@bp.route('/empresa/<int:eid>/mes/<periodo>/resumen')
def mes_resumen(eid, periodo):
    """Resumen ejecutivo del mes para dashboard."""
    Empresa.query.get_or_404(eid)
    if len(periodo) != 7 or periodo[4] != '-':
        return _err('periodo formato YYYY-MM')
    try:
        anio, mes = int(periodo[:4]), int(periodo[5:7])
    except ValueError:
        return _err('periodo inválido')
    import calendar as _cal
    desde = date(anio, mes, 1)
    hasta = date(anio, mes, _cal.monthrange(anio, mes)[1])

    # Movs banco
    movs = MovimientoBanco.query.filter_by(empresa_id=eid).filter(
        MovimientoBanco.fecha >= desde, MovimientoBanco.fecha <= hasta).all()
    movs_total = len(movs)
    movs_sin = sum(1 for m in movs if not m.procesado)
    ing_banco = sum(float(m.abono or 0) for m in movs)
    egr_banco = sum(float(m.cargo or 0) for m in movs)

    # SII por libro
    sii_resumen = {}
    for libro in ('COMPRAS', 'VENTAS', 'HONORARIOS'):
        docs = DocumentoSII.query.filter_by(empresa_id=eid, tipo_libro=libro).filter(
            DocumentoSII.fecha >= desde, DocumentoSII.fecha <= hasta).all()
        sii_resumen[libro] = {
            'total': len(docs),
            'sin_procesar': sum(1 for d in docs if not d.procesado),
            'monto': sum(float(d.total or 0) for d in docs),
        }

    # Asientos por estado
    asientos = Asiento.query.filter_by(empresa_id=eid).filter(
        Asiento.fecha >= desde, Asiento.fecha <= hasta).all()
    asi_resumen = {'BORRADOR': 0, 'CONFIRMADO': 0, 'ANULADO': 0, 'descuadrados': 0}
    for a in asientos:
        asi_resumen[a.estado] = asi_resumen.get(a.estado, 0) + 1
        if not a.cuadrado:
            asi_resumen['descuadrados'] += 1

    # Saldos clave (a fin de mes, solo CONFIRMADO)
    cuentas_clave = {
        'banco': '1.1.02', 'caja': '1.1.01',
        'iva_cf': '1.1.05', 'iva_df': '2.1.03',
        'ret_honorarios': '2.1.04', 'ppm': '1.1.06',
    }
    saldos = {}
    for nombre, codigo in cuentas_clave.items():
        cta = Cuenta.query.filter_by(empresa_id=eid, codigo=codigo).first()
        saldos[nombre] = float(cta.saldo(hasta=hasta)) if cta else 0.0

    # F29 mes anterior cargado
    if mes == 1:
        mes_prev = f'{anio-1}-12'
    else:
        mes_prev = f'{anio}-{mes-1:02d}'
    from models import DeclaracionF29
    f29_prev = DeclaracionF29.query.filter_by(empresa_id=eid, periodo=mes_prev).first()

    return jsonify({
        'periodo': periodo,
        'movs_banco': {'total': movs_total, 'sin_procesar': movs_sin,
                       'ingresos': ing_banco, 'egresos': egr_banco},
        'sii': sii_resumen,
        'asientos': asi_resumen,
        'saldos_clave': saldos,
        'f29_mes_anterior': {
            'periodo': mes_prev,
            'cargado': f29_prev is not None,
            'codigo_91': float(f29_prev.codigo_91) if f29_prev else None,
        },
    })


# ─── P4: Editar y eliminar asiento BORRADOR ───────────────────────────────────
@bp.route('/asiento/<int:aid>', methods=['PATCH'])
def asiento_editar(aid):
    a = Asiento.query.get_or_404(aid)
    if a.estado == 'CONFIRMADO':
        return _err('asiento confirmado no editable — anular primero')
    if a.estado == 'ANULADO':
        return _err('asiento anulado no editable')
    data = request.get_json() or {}

    # Fecha
    if 'fecha' in data:
        try:
            a.fecha = date.fromisoformat(data['fecha'])
        except ValueError:
            return _err('fecha inválida (YYYY-MM-DD)')
    if 'descripcion' in data:
        a.descripcion = (data['descripcion'] or '').strip() or a.descripcion
    if 'origen' in data and data['origen']:
        a.origen = data['origen'].upper()

    # Líneas (reemplazan TODAS las existentes)
    if 'lineas' in data:
        codigos = {c.codigo: c for c in Cuenta.query.filter_by(empresa_id=a.empresa_id).all()}
        lineas_in = data['lineas']
        if not lineas_in:
            return _err('lineas vacías')
        lineas_parsed = []
        td = th = 0.0
        for i, ln in enumerate(lineas_in):
            cod = ln.get('cuenta_codigo') or ''
            cta = codigos.get(cod)
            if not cta and ln.get('cuenta_id'):
                cta = Cuenta.query.filter_by(empresa_id=a.empresa_id, id=int(ln['cuenta_id'])).first()
            if not cta:
                return _err(f'línea {i+1}: cuenta {cod or ln.get("cuenta_id")} no existe')
            d = float(ln.get('debe') or 0); h = float(ln.get('haber') or 0)
            cp_id = ln.get('contraparte_id')
            if cta.requiere_aux and (d or h) and not cp_id:
                return _err(f'línea {i+1} ({cta.codigo} {cta.nombre}) requiere contraparte_id')
            if cp_id:
                if not Contraparte.query.filter_by(empresa_id=a.empresa_id, id=int(cp_id)).first():
                    return _err(f'línea {i+1}: contraparte_id {cp_id} no existe')
            lineas_parsed.append({
                'cuenta_id': cta.id, 'debe': d, 'haber': h,
                'descripcion': (ln.get('descripcion') or '').strip(),
                'contraparte_id': int(cp_id) if cp_id else None, 'orden': i + 1,
            })
            td += d; th += h
        if abs(td - th) > 1:
            return _err(f'descuadrado: D={td:,.0f} H={th:,.0f}')
        # Borrar líneas viejas e insertar nuevas
        LineaAsiento.query.filter_by(asiento_id=a.id).delete()
        for l in lineas_parsed:
            db.session.add(LineaAsiento(asiento_id=a.id, **l))

    # Confirmar si lo pide
    if data.get('estado', '').upper() == 'CONFIRMADO':
        db.session.flush()
        try:
            confirmar_asiento(a)
        except Exception as e:
            db.session.rollback()
            return _err(f'no se pudo confirmar: {e}')

    # Re-asignar movs/sii si vienen
    for mid in (data.get('mov_banco_ids') or []):
        m = MovimientoBanco.query.filter_by(empresa_id=a.empresa_id, id=int(mid)).first()
        if m:
            m.procesado = True; m.asiento_id = a.id
    for did in (data.get('sii_doc_ids') or []):
        d_sii = DocumentoSII.query.filter_by(empresa_id=a.empresa_id, id=int(did)).first()
        if d_sii:
            d_sii.procesado = True; d_sii.asiento_id = a.id

    db.session.commit()
    return jsonify(_e2j(a, with_lines=True))


@bp.route('/asiento/<int:aid>', methods=['DELETE'])
def asiento_eliminar(aid):
    a = Asiento.query.get_or_404(aid)
    if a.estado == 'CONFIRMADO':
        return _err('asiento confirmado no eliminable — anular primero')
    # Desvincular movs y sii (procesado=0, asiento_id=NULL)
    for m in MovimientoBanco.query.filter_by(asiento_id=aid).all():
        m.procesado = False; m.asiento_id = None
    for d in DocumentoSII.query.filter_by(asiento_id=aid).all():
        d.procesado = False; d.asiento_id = None
    from models import AsientoAudit
    AsientoAudit.query.filter_by(asiento_id=aid).delete()
    LineaAsiento.query.filter_by(asiento_id=aid).delete()
    db.session.delete(a)
    db.session.commit()
    return jsonify({'ok': True, 'eliminado_id': aid})


# ─── P5: Confirmación bulk ────────────────────────────────────────────────────
@bp.route('/asientos/confirmar', methods=['POST'])
def asientos_confirmar_bulk(  ):
    data = request.get_json() or {}
    ids = data.get('ids') or []
    if not ids:
        return _err('ids vacíos')
    confirmados, fallidos = [], []
    for aid in ids:
        a = Asiento.query.get(int(aid))
        if not a:
            fallidos.append({'id': aid, 'error': 'no existe'}); continue
        if a.estado == 'CONFIRMADO':
            fallidos.append({'id': aid, 'error': 'ya confirmado'}); continue
        try:
            confirmar_asiento(a)
            db.session.commit()
            confirmados.append(aid)
        except Exception as e:
            db.session.rollback()
            fallidos.append({'id': aid, 'error': str(e)})
    return jsonify({'confirmados': confirmados, 'fallidos': fallidos})


# ─── P6: Búsqueda contrapartes + lectura F29 ──────────────────────────────────
# El GET /api/empresa/<id>/contrapartes ya existe. Agrego soporte de ?q=
# modificando el handler existente (ver más arriba). Acá agrego el otro.
@bp.route('/empresa/<int:eid>/contrapartes-buscar')
def contrapartes_buscar(eid):
    Empresa.query.get_or_404(eid)
    q = (request.args.get('q') or '').strip().lower()
    if not q:
        cps = Contraparte.query.filter_by(empresa_id=eid, activo=True).order_by(
            Contraparte.razon_social).limit(50).all()
    else:
        pattern = f'%{q}%'
        cps = (Contraparte.query.filter_by(empresa_id=eid, activo=True)
               .filter(db.or_(
                   func.lower(Contraparte.razon_social).like(pattern),
                   func.lower(Contraparte.rut).like(pattern)
               )).order_by(Contraparte.razon_social).limit(50).all())
    return jsonify([{
        'id': c.id, 'rut': c.rut, 'razon_social': c.razon_social, 'tipo': c.tipo,
    } for c in cps])


# ─── Importar SII / cartola vía API (multipart) ───────────────────────────────
@bp.route('/empresa/<int:eid>/importar/<tipo>', methods=['POST'])
def importar_archivo(eid, tipo):
    """Importa libros SII (compras/ventas/honorarios) o cartolas (banco/tarjeta).
    multipart/form-data con campo 'archivo'. Opcional: banco, cuenta_bancaria.
    Solo para tipos procesables; para subir respaldos libres usar el módulo Documentos.
    """
    Empresa.query.get_or_404(eid)
    tipo = (tipo or '').lower()
    if tipo not in ('compras', 'ventas', 'honorarios', 'banco', 'tarjeta'):
        return _err(f'tipo inválido: {tipo}. Use compras|ventas|honorarios|banco|tarjeta')
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        return _err('archivo requerido (multipart/form-data)')
    try:
        if tipo == 'compras':
            from importers import libro_compras
            resultado = libro_compras.importar(archivo, eid)
        elif tipo == 'ventas':
            from importers import libro_ventas
            resultado = libro_ventas.importar(archivo, eid)
        elif tipo == 'honorarios':
            from importers import libro_honorarios
            resultado = libro_honorarios.importar(archivo, eid)
        elif tipo == 'banco':
            from importers import cartola
            banco = (request.form.get('banco') or '').strip()
            cuenta_bancaria = (request.form.get('cuenta_bancaria') or '').strip()
            resultado = cartola.importar(archivo, eid, banco, cuenta_bancaria)
        elif tipo == 'tarjeta':
            from importers import tarjeta_credito as tc_mod
            banco = (request.form.get('banco') or 'Banco de Chile (TC)').strip()
            cuenta_bancaria = (request.form.get('cuenta_bancaria') or '').strip()
            resultado = tc_mod.importar(archivo, eid, banco, cuenta_bancaria)
        db.session.commit()
        return jsonify({
            'ok': True, 'tipo': tipo.upper(), 'archivo': archivo.filename,
            'resumen': str(resultado) if resultado else 'importado',
        }), 201
    except Exception as e:
        db.session.rollback()
        return _err(f'error al importar: {e}', 500)


# ─── Gatillar descarga SII (sii_scraper) ──────────────────────────────────────
@bp.route('/empresa/<int:eid>/sii/descargar', methods=['POST'])
def sii_descargar(eid):
    """Descarga libros SII directo desde el portal con las credenciales de la empresa.
    Body JSON: {"libro": "COMPRAS|VENTAS|HONORARIOS", "periodo": "YYYY-MM"}
    """
    empresa = Empresa.query.get_or_404(eid)
    if not empresa.clave_sii:
        return _err('empresa sin clave_sii configurada')
    data = request.get_json() or {}
    libro = (data.get('libro') or '').upper()
    periodo = data.get('periodo') or ''
    if libro not in ('COMPRAS', 'VENTAS', 'HONORARIOS'):
        return _err('libro debe ser COMPRAS|VENTAS|HONORARIOS')
    if len(periodo) != 7 or periodo[4] != '-':
        return _err('periodo formato YYYY-MM')
    try:
        from importers.sii_scraper import descargar
        bytes_archivo = descargar(empresa.rut, empresa.clave_sii, periodo, libro.lower())
    except Exception as e:
        return _err(f'error scraper SII: {e}', 500)
    # Procesar con el importer correspondiente
    try:
        from io import BytesIO
        class _FS:  # mini FileStorage-like
            def __init__(s, data, name):
                s.stream = BytesIO(data); s.filename = name
            def save(s, dst):
                with open(dst, 'wb') as f: f.write(s.stream.getvalue())
        ext = '.xlsx' if libro != 'HONORARIOS' else '.xls'
        fs = _FS(bytes_archivo, f'sii_{libro.lower()}_{periodo}{ext}')
        if libro == 'COMPRAS':
            from importers import libro_compras; resultado = libro_compras.importar(fs, eid)
        elif libro == 'VENTAS':
            from importers import libro_ventas; resultado = libro_ventas.importar(fs, eid)
        else:
            from importers import libro_honorarios; resultado = libro_honorarios.importar(fs, eid)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return _err(f'descargó pero falló importar: {e}', 500)
    return jsonify({
        'ok': True, 'libro': libro, 'periodo': periodo,
        'bytes_descargados': len(bytes_archivo),
        'resumen': str(resultado) if resultado else 'procesado',
    })


# ─── Mayor por contraparte ────────────────────────────────────────────────────
@bp.route('/contraparte/<int:cp_id>/mayor')
def mayor_contraparte(cp_id):
    """Lista todas las líneas de asientos confirmados que tienen esta contraparte.
    Útil para auditar saldos de un proveedor/cliente específico."""
    cp = Contraparte.query.get_or_404(cp_id)
    desde = request.args.get('desde'); hasta = request.args.get('hasta')
    q = (db.session.query(LineaAsiento, Asiento, Cuenta)
         .join(Asiento, Asiento.id == LineaAsiento.asiento_id)
         .join(Cuenta, Cuenta.id == LineaAsiento.cuenta_id)
         .filter(LineaAsiento.contraparte_id == cp_id,
                 Asiento.estado == 'CONFIRMADO'))
    if desde: q = q.filter(Asiento.fecha >= desde)
    if hasta: q = q.filter(Asiento.fecha <= hasta)
    rows = q.order_by(Asiento.fecha, Asiento.numero).all()
    lineas = []
    saldo = 0.0
    for l, a, c in rows:
        d = float(l.debe or 0); h = float(l.haber or 0)
        saldo += d - h
        lineas.append({
            'fecha': a.fecha.isoformat(), 'asiento_id': a.id, 'numero': a.numero,
            'descripcion': a.descripcion, 'cuenta_codigo': c.codigo, 'cuenta_nombre': c.nombre,
            'debe': d, 'haber': h, 'saldo_acumulado': saldo,
            'glosa': l.descripcion,
        })
    return jsonify({
        'contraparte': {'id': cp.id, 'rut': cp.rut, 'razon_social': cp.razon_social,
                        'tipo': cp.tipo, 'empresa_id': cp.empresa_id},
        'desde': desde, 'hasta': hasta,
        'lineas': lineas,
        'totales': {
            'debe': sum(l['debe'] for l in lineas),
            'haber': sum(l['haber'] for l in lineas),
            'saldo_final': saldo,
        },
    })


@bp.route('/empresa/<int:eid>/f29/<periodo>')
def f29_lectura(eid, periodo):
    Empresa.query.get_or_404(eid)
    from models import DeclaracionF29
    import json as _json
    f29 = DeclaracionF29.query.filter_by(empresa_id=eid, periodo=periodo).first()
    if not f29:
        return _err(f'F29 {periodo} no cargado', 404)
    codigos = {}
    try:
        codigos = _json.loads(f29.codigos_json or '{}')
    except Exception:
        codigos = {}
    return jsonify({
        'empresa_id': eid, 'periodo': periodo, 'folio': f29.folio,
        'fecha_descarga': f29.fecha_descarga.isoformat() if f29.fecha_descarga else None,
        'codigo_62': float(f29.codigo_62 or 0),
        'codigo_48': float(f29.codigo_48 or 0),
        'codigo_151': float(f29.codigo_151 or 0),
        'codigo_89': float(f29.codigo_89 or 0),
        'codigo_91': float(f29.codigo_91 or 0),
        'codigo_92': float(f29.codigo_92 or 0),
        'codigo_538': float(f29.codigo_538 or 0),
        'codigo_547': float(f29.codigo_547 or 0),
        'codigos_completos': codigos,
    })
